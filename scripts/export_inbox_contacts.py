#!/usr/bin/env python3
"""export_inbox_contacts.py -- the whole directory in one spreadsheet. READ-ONLY.

A bird's-eye view of every contact, every field, plus a state-of-the-data
summary so you can see exactly where you are and where the gaps are.

Tabs:
  1. Overview   -- totals, category/priority/vertical breakdowns, and a
                   "data completeness" table (how many are missing name / email
                   / role / LinkedIn / phone / org).
  2. Contacts   -- one row per inbox contact, all fields, role-dictionary
                   vertical joined in, frozen header + autofilter on every col.

Lead-generator contacts live in a separate table (lead_contacts); pass
--with-leads to add a third tab for them too.

Usage (repo root, venv active):
  python scripts/export_inbox_contacts.py
  python scripts/export_inbox_contacts.py --out C:\\Users\\it2\\Desktop\\contacts.xlsx
  python scripts/export_inbox_contacts.py --with-leads
  python scripts/export_inbox_contacts.py --include-junk      # default hides junk

Falls back to CSV (Contacts only) if openpyxl isn't installed.
"""

import argparse
import asyncio
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

try:
    from dotenv import load_dotenv

    load_dotenv()
except ImportError:
    pass

from sqlalchemy import text  # noqa: E402

from app.database import async_session  # noqa: E402

CONTACTS_SQL = """
    SELECT c.id, c.display_name, c.first_name, c.last_name, c.email,
           c.organization, c.parent_company, c.management_company,
           COALESCE(NULLIF(c.title,''), c.inferred_role) AS role,
           c.seniority, c.department, c.contact_category, c.procurement_priority,
           r.vertical AS dict_vertical,
           c.is_decision_maker, c.brand_tier, c.gpo,
           c.linkedin_url, c.phone, c.address,
           c.interaction_count, c.relevance_verdict, c.relevance_score,
           c.enrichment_source, c.enrichment_confidence,
           to_char(c.enriched_at,'YYYY-MM-DD') AS enriched_at,
           c.org_source, c.approval_status,
           c.matched_hotel_id, c.matched_lead_id,
           array_to_string(c.source_mailboxes, '; ') AS mailboxes,
           to_char(c.first_seen,'YYYY-MM-DD') AS first_seen,
           to_char(c.last_seen,'YYYY-MM-DD') AS last_seen,
           COALESCE(c.is_shared_mailbox,false) AS shared
    FROM contacts c
    LEFT JOIN contact_roles r
      ON r.role_normalized = lower(regexp_replace(
             regexp_replace(COALESCE(NULLIF(c.title,''), c.inferred_role, ''),
             '&', ' and ', 'g'), '[^a-zA-Z0-9 ]+', ' ', 'g'))
    {where}
    ORDER BY
      CASE c.contact_category WHEN 'buyer' THEN 0 WHEN 'seller' THEN 2
        WHEN 'competitor' THEN 3 WHEN 'personal' THEN 4 WHEN 'junk' THEN 5
        ELSE 1 END,
      CASE c.procurement_priority WHEN 'P1' THEN 0 WHEN 'P2' THEN 1
        WHEN 'P3' THEN 2 WHEN 'P4' THEN 3 ELSE 4 END,
      c.organization, c.id
"""

COLS = [
    ("id", "id", 7), ("display_name", "Name", 24), ("email", "Email", 30),
    ("organization", "Organization", 26), ("role", "Role", 26),
    ("seniority", "Seniority", 11), ("contact_category", "Category", 11),
    ("procurement_priority", "Priority", 9), ("dict_vertical", "Vertical", 13),
    ("is_decision_maker", "DM", 4), ("brand_tier", "Brand tier", 16),
    ("gpo", "GPO", 10), ("linkedin_url", "LinkedIn", 30), ("phone", "Phone", 16),
    ("interaction_count", "Emails", 7), ("relevance_verdict", "Relevance", 10),
    ("enrichment_source", "Enriched via", 12), ("enriched_at", "Enriched", 11),
    ("parent_company", "Parent co", 22), ("management_company", "Mgmt co", 22),
    ("department", "Dept", 16), ("org_source", "Org source", 13),
    ("approval_status", "Status", 14), ("mailboxes", "Source mailboxes", 30),
    ("first_seen", "First seen", 11), ("last_seen", "Last seen", 11),
]


async def fetch(session, include_junk):
    where = "" if include_junk else "WHERE (c.contact_category IS NULL OR c.contact_category <> 'junk')"
    return (await session.execute(text(CONTACTS_SQL.format(where=where)))).all()


def completeness(rows):
    n = len(rows) or 1
    def miss(pred):
        c = sum(1 for r in rows if pred(r))
        return c, round(100 * c / n)
    return {
        "Total contacts": (len(rows), 100),
        "Missing real name": miss(lambda r: not (r.first_name or r.last_name)),
        "Missing email": miss(lambda r: not r.email),
        "Missing role/title": miss(lambda r: not r.role),
        "Missing LinkedIn": miss(lambda r: not r.linkedin_url),
        "Missing phone": miss(lambda r: not r.phone),
        "Missing organization": miss(lambda r: not r.organization),
        "No priority (P_unknown)": miss(lambda r: (r.procurement_priority or "P_unknown") == "P_unknown"),
        "Uncategorized": miss(lambda r: not r.contact_category),
    }


def breakdown(rows, key, order=None):
    from collections import Counter
    c = Counter(getattr(r, key) or "(none)" for r in rows)
    items = sorted(c.items(), key=lambda kv: (order.index(kv[0]) if order and kv[0] in order else 99, -kv[1]))
    return items


def write_xlsx(rows, out_path, with_leads, leads):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return False

    wb = Workbook()
    # ---- Overview tab ----
    ov = wb.active
    ov.title = "Overview"
    hdr = Font(name="Arial", bold=True, size=13)
    sub = Font(name="Arial", bold=True, size=11, color="1F2A44")
    norm = Font(name="Arial", size=10)
    ov["A1"] = "JA Uniforms -- Contact Directory snapshot"
    ov["A1"].font = hdr
    ov.merge_cells("A1:C1")

    row = 3
    ov.cell(row, 1, "DATA COMPLETENESS").font = sub
    row += 1
    ov.cell(row, 1, "Metric").font = Font(bold=True, size=10)
    ov.cell(row, 2, "Count").font = Font(bold=True, size=10)
    ov.cell(row, 3, "% of total").font = Font(bold=True, size=10)
    row += 1
    for label, (cnt, pct) in completeness(rows).items():
        ov.cell(row, 1, label).font = norm
        ov.cell(row, 2, cnt).font = norm
        ov.cell(row, 3, f"{pct}%").font = norm
        if "Missing" in label or "Uncategorized" in label or "No priority" in label:
            shade = "FDECEA" if pct >= 40 else ("FFF7E6" if pct >= 15 else "EAF6EC")
            for col in (1, 2, 3):
                ov.cell(row, col).fill = PatternFill("solid", fgColor=shade)
        row += 1

    row += 1
    ov.cell(row, 1, "BY CATEGORY").font = sub
    row += 1
    for k, v in breakdown(rows, "contact_category",
                          ["buyer", "seller", "competitor", "personal", "(none)"]):
        ov.cell(row, 1, k).font = norm
        ov.cell(row, 2, v).font = norm
        row += 1
    row += 1
    ov.cell(row, 1, "BY PRIORITY").font = sub
    row += 1
    for k, v in breakdown(rows, "procurement_priority", ["P1", "P2", "P3", "P4", "P_unknown"]):
        ov.cell(row, 1, k).font = norm
        ov.cell(row, 2, v).font = norm
        row += 1
    row += 1
    ov.cell(row, 1, "BY VERTICAL (role dictionary)").font = sub
    row += 1
    for k, v in breakdown(rows, "dict_vertical"):
        ov.cell(row, 1, k).font = norm
        ov.cell(row, 2, v).font = norm
        row += 1
    ov.column_dimensions["A"].width = 32
    ov.column_dimensions["B"].width = 12
    ov.column_dimensions["C"].width = 12

    # ---- Contacts tab ----
    _sheet(wb, "Contacts", rows)

    # ---- Leads tab ----
    if with_leads and leads:
        _leads_sheet(wb, leads)

    wb.save(out_path)
    return True


def _sheet(wb, title, rows):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet(title)
    hdr_fill = PatternFill("solid", fgColor="1F2A44")
    thin = Side(style="thin", color="E0E0E0")
    border = Border(bottom=thin)
    pri_fill = {"P1": "C6EFCE", "P2": "FFF2CC", "P3": "DDEBF7", "P4": "F2F2F2"}
    for ci, (_, label, _w) in enumerate(COLS, 1):
        cell = ws.cell(1, ci, label)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = hdr_fill
        cell.alignment = Alignment(vertical="center")
    for ri, r in enumerate(rows, 2):
        pri = r.procurement_priority or ""
        for ci, (attr, _label, _w) in enumerate(COLS, 1):
            v = getattr(r, attr, None)
            if isinstance(v, bool):
                v = "yes" if v else ""
            cell = ws.cell(ri, ci, v)
            cell.font = Font(name="Arial", size=9)
            cell.border = border
            if attr == "procurement_priority" and pri in pri_fill:
                cell.fill = PatternFill("solid", fgColor=pri_fill[pri])
    for ci, (_, _label, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(rows)+1}"


def _leads_sheet(wb, leads):
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet("Lead generator")
    cols = [("name", "Name", 24), ("organization", "Organization", 26),
            ("title", "Title", 26), ("email", "Email", 30), ("linkedin", "LinkedIn", 30),
            ("phone", "Phone", 16), ("tier", "Tier", 10), ("score", "Score", 8),
            ("is_saved", "Saved", 7), ("found_via", "Found via", 16)]
    for ci, (_, label, _w) in enumerate(cols, 1):
        cell = ws.cell(1, ci, label)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1F2A44")
    for ri, r in enumerate(leads, 2):
        for ci, (attr, _l, _w) in enumerate(cols, 1):
            v = getattr(r, attr, None)
            if isinstance(v, bool):
                v = "yes" if v else ""
            ws.cell(ri, ci, v).font = Font(name="Arial", size=9)
    for ci, (_, _l, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(leads)+1}"


def write_csv(rows, out_path):
    import csv
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow([label for _, label, _ in COLS])
        for r in rows:
            line = []
            for attr, _l, _w in COLS:
                v = getattr(r, attr, None)
                line.append("yes" if v is True else ("" if v is False else v))
            w.writerow(line)


async def run(out, with_leads, include_junk):
    async with async_session() as session:
        rows = await fetch(session, include_junk)
        leads = []
        if with_leads:
            try:
                leads = (await session.execute(text(
                    "SELECT name, organization, title, email, linkedin, phone, "
                    "tier, score, is_saved, found_via FROM lead_contacts "
                    "ORDER BY is_saved DESC, score DESC NULLS LAST LIMIT 20000"
                ))).all()
            except Exception as e:
                print(f"(could not load lead_contacts: {e})")

    print(f"Loaded {len(rows):,} contacts" + (f" + {len(leads):,} leads" if leads else ""))
    target = Path(out)
    if write_xlsx(rows, str(target), with_leads, leads):
        print(f"\nWrote {target}")
        print("  Tab 1 'Overview'  -- totals, completeness, category/priority/vertical")
        print("  Tab 2 'Contacts'  -- every contact, all fields, filterable")
        if with_leads and leads:
            print("  Tab 3 'Lead generator' -- discovery leads")
    else:
        csv_path = target.with_suffix(".csv")
        write_csv(rows, str(csv_path))
        print(f"\nopenpyxl missing; wrote CSV -> {csv_path}")
        print("For the multi-tab Excel:  pip install openpyxl  then re-run.")


def main():
    p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    p.add_argument("--out", default="inbox_contacts_export.xlsx")
    p.add_argument("--with-leads", action="store_true", help="add a lead-generator tab")
    p.add_argument("--include-junk", action="store_true", help="include junk contacts")
    args = p.parse_args()
    asyncio.run(run(args.out, args.with_leads, args.include_junk))


if __name__ == "__main__":
    main()
