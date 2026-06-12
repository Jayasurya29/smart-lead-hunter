#!/usr/bin/env python3
"""export_apply_preview.py -- the 639 contacts as a real Excel sheet. READ-ONLY.

Runs the SAME query --apply uses and writes a formatted .xlsx you open in Excel:
priority-colored rows, grouped P1 -> P4, a filter on every column, and a summary
row up top. Nothing in the database is changed.

    python scripts/export_apply_preview.py
    python scripts/export_apply_preview.py --out C:\\Users\\it2\\Desktop\\priority_preview.xlsx

If openpyxl isn't installed it falls back to a .csv automatically (and tells you
the one-line pip command to get the pretty version).
"""

import argparse
import asyncio
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

try:
    from dotenv import load_dotenv

    load_dotenv()
except ImportError:
    pass

from sqlalchemy import text  # noqa: E402

from app.database import async_session  # noqa: E402

SEL = text(
    """
    SELECT c.id,
           COALESCE(c.display_name, NULLIF(trim(c.first_name||' '||c.last_name),''),
                    c.email) AS who,
           COALESCE(c.email,'') AS email,
           COALESCE(c.organization,'') AS org,
           COALESCE(NULLIF(c.title,''), NULLIF(c.inferred_role,'')) AS role,
           COALESCE(c.contact_category,'') AS cat,
           r.priority AS new_pri,
           r.vertical AS vertical,
           r.is_relevant AS relevant,
           COALESCE(c.procurement_priority,'P_unknown') AS old_pri
    FROM contacts c
    JOIN contact_roles r
      ON r.role_normalized = lower(regexp_replace(
             regexp_replace(COALESCE(NULLIF(c.title,''), NULLIF(c.inferred_role,'')),
             '&', ' and ', 'g'), '[^a-zA-Z0-9 ]+', ' ', 'g'))
    WHERE (c.procurement_priority IS NULL OR c.procurement_priority = 'P_unknown')
      AND r.priority <> 'P_unknown'
      AND (c.contact_category IS NULL OR c.contact_category NOT IN
           ('junk','seller','competitor','personal','operational'))
      AND COALESCE(c.is_shared_mailbox,false) = false
    ORDER BY CASE r.priority WHEN 'P1' THEN 0 WHEN 'P2' THEN 1
             WHEN 'P3' THEN 2 ELSE 3 END,
             (c.contact_category='buyer') DESC, c.organization, c.id
    """
)

HEADERS = ["ID", "Name", "Email", "Organization", "Role (matched)",
           "Category", "New Priority", "Vertical", "Relevant", "Current"]

# priority -> (row fill, font color)
PRI_FILL = {
    "P1": ("C6EFCE", "06430F"),   # green  -- top buyers
    "P2": ("FFF2CC", "7F6000"),   # amber  -- decision-makers
    "P3": ("DDEBF7", "1F4E79"),   # blue   -- secondary
    "P4": ("F2F2F2", "808080"),   # gray   -- low / non-buyer
}


def _write_xlsx(rows, out_path: str) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Priority Preview"

    counts = {}
    for r in rows:
        counts[r.new_pri] = counts.get(r.new_pri, 0) + 1
    summary = (f"{len(rows)} contacts would be re-prioritized   "
               f"P1={counts.get('P1',0)}  P2={counts.get('P2',0)}  "
               f"P3={counts.get('P3',0)}  P4={counts.get('P4',0)}   "
               f"(READ-ONLY preview -- nothing applied yet)")
    ws["A1"] = summary
    ws["A1"].font = Font(name="Arial", bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws.row_dimensions[1].height = 22

    hdr_fill = PatternFill("solid", fgColor="1F2A44")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, name in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=c, value=name)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

    for i, r in enumerate(rows, start=3):
        vals = [r.id, r.who, r.email, r.org, r.role, r.cat, r.new_pri,
                r.vertical, ("yes" if r.relevant else ("no" if r.relevant is False else "?")),
                r.old_pri]
        fill_hex, font_hex = PRI_FILL.get(r.new_pri, ("FFFFFF", "000000"))
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = Font(name="Arial", size=10, color=font_hex,
                             bold=(c == 7))  # bold the priority column
            cell.fill = PatternFill("solid", fgColor=fill_hex)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border

    widths = [7, 26, 30, 30, 30, 11, 12, 15, 9, 10]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{len(rows)+2}"

    wb.save(out_path)
    return True


def _write_csv(rows, out_path: str) -> None:
    import csv
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        for r in rows:
            w.writerow([r.id, r.who, r.email, r.org, r.role, r.cat, r.new_pri,
                        r.vertical,
                        "yes" if r.relevant else ("no" if r.relevant is False else "?"),
                        r.old_pri])


async def run(out: str) -> None:
    async with async_session() as session:
        rows = (await session.execute(SEL)).all()

    if not rows:
        print("No contacts match -- nothing to preview.")
        return

    counts = {}
    for r in rows:
        counts[r.new_pri] = counts.get(r.new_pri, 0) + 1
    print(f"{len(rows)} contacts:  " +
          "  ".join(f"{k}={counts.get(k,0)}" for k in ("P1", "P2", "P3", "P4")))

    target = Path(out)
    if _write_xlsx(rows, str(target)):
        print(f"\nWrote formatted Excel sheet -> {target}")
        print("Open it in Excel: rows are colored by priority (green=P1 buyers, "
              "amber=P2 decision-makers), every column filterable.")
    else:
        csv_path = target.with_suffix(".csv")
        _write_csv(rows, str(csv_path))
        print(f"\nopenpyxl not installed, wrote CSV instead -> {csv_path}")
        print("For the colored Excel version:  pip install openpyxl  then re-run.")
    print("\nNothing was changed. When the list looks right, apply for real:")
    print("    python scripts/build_role_dictionary.py --apply")


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    p.add_argument("--out", type=str, default="priority_preview.xlsx",
                   help="Output path (.xlsx). Default: priority_preview.xlsx in cwd")
    args = p.parse_args()
    asyncio.run(run(args.out))


if __name__ == "__main__":
    main()
