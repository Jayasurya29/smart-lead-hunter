"""
Export contacts for existing hotels — polished 2-sheet Excel workbook.

Sheet 1: Hotel Call List (one row per hotel, primary contact, revenue)
Sheet 2: All Contacts (one row per contact, every useful field)

Usage:
    python scripts/export_contacts.py                         # all existing hotels
    python scripts/export_contacts.py --source towne_park     # filter by data_source
    python scripts/export_contacts.py --source towne_park -o TownePark_Contacts.xlsx
"""

import asyncio
import argparse
import sys
import os
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from dotenv import load_dotenv
load_dotenv()


async def export(source: str = None, output: str = None):
    from app.database import async_session
    from sqlalchemy import text

    src_filter = f"WHERE eh.data_source = '{source}'" if source else ""
    src_and = f"AND eh.data_source = '{source}'" if source else ""

    async with async_session() as s:
        # ── Hotels ──
        result = await s.execute(text(f"""
            SELECT eh.id, eh.hotel_name, eh.brand, eh.brand_tier, eh.city, eh.state,
                   eh.zone, eh.address, eh.zip_code, eh.opening_date, eh.hotel_type,
                   eh.owner, eh.management_company, eh.developer, eh.lead_score,
                   eh.room_count, eh.revenue_opening, eh.revenue_annual,
                   eh.is_client, eh.status, eh.created_at, eh.hotel_website,
                   eh.country, eh.description
            FROM existing_hotels eh
            {src_filter}
            ORDER BY eh.lead_score DESC NULLS LAST
        """))
        hotels = result.all()

        # ── All contacts ──
        result = await s.execute(text(f"""
            SELECT lc.id, lc.existing_hotel_id, eh.hotel_name, eh.brand, eh.city,
                   eh.state, eh.brand_tier, eh.room_count, eh.management_company,
                   lc.name, lc.title, lc.organization, lc.email, lc.secondary_email,
                   lc.phone, lc.linkedin, lc.scope, lc.score, lc.strategist_priority,
                   lc.is_primary, lc.is_saved, lc.found_via, lc.confidence, lc.tier,
                   lc.created_at, lc.source_detail, lc.strategist_reasoning,
                   lc.evidence_url
            FROM lead_contacts lc
            JOIN existing_hotels eh ON eh.id = lc.existing_hotel_id
            WHERE lc.name IS NOT NULL {src_and}
            ORDER BY eh.hotel_name, lc.is_primary DESC, lc.score DESC NULLS LAST
        """))
        contacts = result.all()

    if not hotels:
        print(f"No hotels found{' for source=' + source if source else ''}.")
        return

    # Contact lookup
    contacts_by_hotel = {}
    for c in contacts:
        contacts_by_hotel.setdefault(c[1], []).append(c)

    print(f"Found {len(hotels)} hotels with {len(contacts)} total contacts.")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Palette ──
    NAVY = PatternFill(start_color="0F1D32", fill_type="solid")
    SUBNAV = PatternFill(start_color="1E293B", fill_type="solid")
    ZEBRA = PatternFill(start_color="F8FAFC", fill_type="solid")
    GOLD_BG = PatternFill(start_color="FEF9EF", fill_type="solid")
    GREEN_BG = PatternFill(start_color="ECFDF5", fill_type="solid")
    SECTION = PatternFill(start_color="EFF6FF", fill_type="solid")
    HDR = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    Font(name="Calibri", size=10, bold=True, color="1E40AF")
    D = Font(name="Calibri", size=10)
    B = Font(name="Calibri", size=10, bold=True)
    LINK = Font(name="Calibri", size=10, color="3B82F6", underline="single")
    GREEN = Font(name="Calibri", size=10, bold=True, color="059669")
    AMBER = Font(name="Calibri", size=10, bold=True, color="D97706")
    RED = Font(name="Calibri", size=10, bold=True, color="DC2626")
    TITLE = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    SUB = Font(name="Calibri", size=10, italic=True, color="94A3B8")
    BORDER = Border(bottom=Side(style="thin", color="E2E8F0"))

    src_label = source.replace("_", " ").title() if source else "All Hotels"
    today = date.today().isoformat()

    def _banner(ws, title_text, subtitle_text, last_col):
        lc = get_column_letter(last_col)
        ws.merge_cells(f"A1:{lc}1")
        ws["A1"].value = title_text
        ws["A1"].font = TITLE
        ws["A1"].fill = NAVY
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36
        ws.merge_cells(f"A2:{lc}2")
        ws["A2"].value = subtitle_text
        ws["A2"].font = SUB
        ws["A2"].fill = SUBNAV
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 22

    def _headers(ws, row, headers, widths):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = HDR
            cell.fill = NAVY
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 32
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = w

    def _write_row(ws, row, values, row_idx, hyperlinks=None):
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = D
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if row_idx % 2 == 1:
                cell.fill = ZEBRA
        if hyperlinks:
            for col, url in hyperlinks.items():
                if url:
                    cell = ws.cell(row=row, column=col)
                    cell.font = LINK
                    try:
                        if url.startswith("mailto:"):
                            cell.hyperlink = url
                        else:
                            cell.hyperlink = url
                    except Exception:
                        pass

    # ════════════════════════════════════════════════════════════
    # SHEET 1: HOTEL CALL LIST
    # ════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = f"{src_label} Hotels"

    h1 = [
        "ID", "Hotel Name", "Brand", "Tier", "City", "State", "Zone",
        "Address", "Zip", "Opening", "Type", "Rooms", "Owner",
        "Management Co", "Developer", "Score",
        "Revenue (Opening)", "Revenue (Annual)", "# Contacts",
        "Primary Contact", "Contact Title", "Contact Email",
        "Contact Phone", "Website", "Client?", "Status", "Added",
    ]
    w1 = [
        6, 38, 18, 18, 14, 10, 13,
        35, 8, 14, 10, 7, 25,
        25, 25, 7,
        15, 15, 10,
        22, 28, 26,
        14, 30, 8, 8, 11,
    ]

    _banner(ws1, f"JA Uniforms — {src_label} Call List",
            f"{today}  •  {len(hotels)} hotels  •  {len(contacts)} contacts", len(h1))
    _headers(ws1, 3, h1, w1)
    ws1.auto_filter.ref = f"A3:{get_column_letter(len(h1))}{3 + len(hotels)}"
    ws1.freeze_panes = "A4"

    clients = 0
    w_contacts = 0

    for idx, h in enumerate(hotels):
        r = idx + 4
        (hid, name, brand, tier, city, state, zone, address, zipcode,
         opening, htype, owner, mgmt, developer, score, rooms,
         rev_o, rev_a, is_client, status, created, website, country, desc) = h

        hc = contacts_by_hotel.get(hid, [])
        pri = next((c for c in hc if c[19]), None) or (hc[0] if hc else None)

        if hc: w_contacts += 1
        if is_client: clients += 1

        tier_lbl = (tier or "").replace("tier1_", "T1 ").replace("tier2_", "T2 ").replace(
            "tier3_", "T3 ").replace("tier4_", "T4 ").replace("tier5_", "T5 ").replace("_", " ").title()

        vals = [
            hid, name, brand or "", tier_lbl, city or "", state or "", zone or "",
            address or "", zipcode or "", opening or "", htype or "", rooms or "",
            owner or "", mgmt or "", developer or "", score or "",
            rev_o, rev_a, len(hc),
            pri[9] if pri else "", pri[10] if pri else "", pri[12] if pri else "",
            pri[14] if pri else "", website or "",
            "Client" if is_client else "Prospect", status or "",
            str(created)[:10] if created else "",
        ]

        links = {}
        if pri and pri[12]:  # email
            links[22] = f"mailto:{pri[12]}"
        if website:
            links[24] = website

        _write_row(ws1, r, vals, idx, links)

        # Score color
        sc = ws1.cell(row=r, column=16)
        if score and score >= 80: sc.font = GREEN
        elif score and score >= 60: sc.font = AMBER
        elif score and score < 40: sc.font = RED

        # Revenue formatting
        for c in (17, 18):
            cell = ws1.cell(row=r, column=c)
            if cell.value:
                cell.number_format = '$#,##0'

    # Summary footer
    sr = len(hotels) + 5
    for label, val in [("SUMMARY", ""), ("Total Hotels:", len(hotels)),
                       ("With Contacts:", w_contacts),
                       ("Zero Contacts:", len(hotels) - w_contacts),
                       ("Clients:", clients)]:
        ws1.cell(row=sr, column=1, value=label).font = B
        ws1.cell(row=sr, column=2, value=val).font = B
        sr += 1

    # ════════════════════════════════════════════════════════════
    # SHEET 2: ALL CONTACTS
    # ════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("All Contacts")

    h2 = [
        "Hotel ID", "Hotel Name", "Brand", "Tier", "City", "State", "Rooms",
        "Mgmt Company",
        "Contact Name", "Title", "Organization",
        "Email", "Secondary Email", "Phone", "LinkedIn",
        "Scope", "Score", "Priority", "Confidence",
        "Primary?", "Saved?",
        "Source Detail", "Found Via", "Evidence URL", "Added",
    ]
    w2 = [
        7, 36, 16, 16, 13, 10, 7,
        22,
        24, 30, 24,
        26, 26, 14, 45,
        16, 7, 7, 10,
        8, 7,
        35, 13, 30, 11,
    ]

    emails_found = 0
    sec_emails_found = 0
    li_found = 0

    _banner(ws2, f"JA Uniforms — {src_label} All Contacts",
            f"{today}  •  {len(contacts)} contacts across {len(hotels)} hotels", len(h2))
    _headers(ws2, 3, h2, w2)
    ws2.auto_filter.ref = f"A3:{get_column_letter(len(h2))}{3 + len(contacts)}"
    ws2.freeze_panes = "I4"  # Freeze hotel cols, scroll contacts

    for idx, c in enumerate(contacts):
        r = idx + 4
        (cid, hotel_id, hotel_name, brand, city, state, tier, rooms, mgmt,
         name, title, org, email, sec_email, phone, linkedin,
         scope, score, priority, is_primary, is_saved, found_via,
         confidence, contact_tier, created, source_detail,
         strategist_reasoning, evidence_url) = c

        if email: emails_found += 1
        if sec_email: sec_emails_found += 1
        if linkedin: li_found += 1

        tier_lbl = (tier or "").replace("tier1_", "T1 ").replace("tier2_", "T2 ").replace(
            "tier3_", "T3 ").replace("tier4_", "T4 ").replace("tier5_", "T5 ").replace("_", " ").title()
        scope_lbl = (scope or "unknown").replace("_", " ").title()

        vals = [
            hotel_id, hotel_name, brand or "", tier_lbl, city or "", state or "",
            rooms or "", mgmt or "",
            name, title or "", org or "",
            email or "", sec_email or "", phone or "", linkedin or "",
            scope_lbl, score or "", priority or "", (confidence or "").title(),
            "Yes" if is_primary else "", "Yes" if is_saved else "",
            source_detail or "", found_via or "", evidence_url or "",
            str(created)[:10] if created else "",
        ]

        links = {}
        if email: links[12] = f"mailto:{email}"
        if sec_email: links[13] = f"mailto:{sec_email}"
        if linkedin: links[15] = linkedin
        if evidence_url: links[24] = evidence_url

        _write_row(ws2, r, vals, idx, links)

        # Primary = bold green name
        if is_primary:
            ws2.cell(row=r, column=9).font = GREEN

        # Priority color
        pri_cell = ws2.cell(row=r, column=18)
        if priority == "P1": pri_cell.font = GREEN
        elif priority == "P2": pri_cell.font = AMBER
        elif priority == "P3": pri_cell.font = D
        elif priority == "P4": pri_cell.font = RED

        # Score color
        sc_cell = ws2.cell(row=r, column=17)
        if score and score >= 20: sc_cell.font = GREEN
        elif score and score >= 10: sc_cell.font = AMBER
        elif score and score < 10: sc_cell.font = RED

        # Scope section coloring for hotel_specific
        if scope and "hotel_specific" in scope.lower():
            ws2.cell(row=r, column=16).fill = GREEN_BG
        elif scope and "owner" in scope.lower():
            ws2.cell(row=r, column=16).fill = GOLD_BG

    # ════════════════════════════════════════════════════════════
    # SHEET 3: SUMMARY
    # ════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Summary")
    _banner(ws3, f"{src_label} Export Summary", today, 4)

    # Coverage stats
    data = [
        ("COVERAGE", "", "", ""),
        ("Total Hotels", len(hotels), "Total Contacts", len(contacts)),
        ("Hotels with Contacts", w_contacts, "Avg Contacts/Hotel",
         f"{len(contacts)/len(hotels):.1f}" if hotels else "0"),
        ("Hotels without Contacts", len(hotels) - w_contacts, "Clients", clients),
        ("", "", "", ""),
        ("CONTACT DATA QUALITY", "", "", ""),
        ("With Primary Email", emails_found,
         "Coverage", f"{emails_found/len(contacts)*100:.0f}%" if contacts else "0%"),
        ("With Secondary Email", sec_emails_found,
         "Coverage", f"{sec_emails_found/len(contacts)*100:.0f}%" if contacts else "0%"),
        ("With LinkedIn", li_found,
         "Coverage", f"{li_found/len(contacts)*100:.0f}%" if contacts else "0%"),
        ("With Phone", sum(1 for c in contacts if c[14]),
         "Coverage", f"{sum(1 for c in contacts if c[14])/len(contacts)*100:.0f}%" if contacts else "0%"),
    ]

    # Scope breakdown
    scope_counts = {}
    priority_counts = {}
    for c in contacts:
        s = (c[16] or "unknown").replace("_", " ").title()
        scope_counts[s] = scope_counts.get(s, 0) + 1
        p = c[18] or "None"
        priority_counts[p] = priority_counts.get(p, 0) + 1

    data.append(("", "", "", ""))
    data.append(("BY SCOPE", "Count", "BY PRIORITY", "Count"))
    scope_items = sorted(scope_counts.items(), key=lambda x: -x[1])
    pri_items = sorted(priority_counts.items())
    max_rows = max(len(scope_items), len(pri_items))
    for i in range(max_rows):
        sl, sc = scope_items[i] if i < len(scope_items) else ("", "")
        pl, pc = pri_items[i] if i < len(pri_items) else ("", "")
        data.append((sl, sc, pl, pc))

    for i, (a, b, c_val, d_val) in enumerate(data, 3):
        for col, val in enumerate([a, b, c_val, d_val], 1):
            cell = ws3.cell(row=i, column=col, value=val)
            cell.font = D
            cell.alignment = Alignment(vertical="center")
        # Section headers bold
        if isinstance(a, str) and a.isupper() and a:
            for col in range(1, 5):
                ws3.cell(row=i, column=col).font = B
                ws3.cell(row=i, column=col).fill = SECTION

    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 28
    ws3.column_dimensions["D"].width = 14

    # ── Save ──
    if not output:
        output = f"{src_label.replace(' ', '')}_{len(hotels)}_Hotels_Export.xlsx"
    wb.save(output)
    print(f"\nExported to: {output}")
    print(f"  Sheet 1: {len(hotels)} hotels (call list with primary contact)")
    print(f"  Sheet 2: {len(contacts)} contacts (all details)")
    print("  Sheet 3: Summary dashboard")
    print(f"  Emails: {emails_found} primary + {sec_emails_found} secondary")
    print(f"  LinkedIn: {li_found}")


def main():
    parser = argparse.ArgumentParser(description="Export existing hotel contacts to Excel")
    parser.add_argument("--source", help="Filter by data_source (e.g. towne_park)")
    parser.add_argument("--output", "-o", help="Output filename")
    args = parser.parse_args()
    asyncio.run(export(source=args.source, output=args.output))


if __name__ == "__main__":
    main()
