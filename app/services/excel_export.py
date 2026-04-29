"""
Shared Excel export builder — JA Uniforms Lead Generator
=========================================================

Used by both /leads/export (New Hotels) and /existing-hotels/export
(Existing Hotels) to produce visually-polished workbooks. Two sheets:

  Sheet 1 — Call List
      One row per hotel, optimized for sales outreach. Color-coded
      score badges, tier chips, timeline badges (new-hotels only),
      hyperlinks on website + LinkedIn, autofilter dropdowns, frozen
      header. Sales can work directly in Excel.

  Sheet 2 — Summary / Score Distribution
      Pivot-friendly aggregation tables. Score buckets, tier breakdown,
      state/zone leaderboards, contact-coverage health. Boss view.

Design palette (matches the JA Uniforms web UI):
  Navy        #0F1D32  primary header bar
  Gold        #D4A853  Tier 1 / Ultra Luxury
  Silver      #C0C0C0  Tier 2 / Luxury
  Bronze      #B07333  Tier 3 / Upper Upscale
  Stone       #6B665E  Tier 4 / Upscale
  Emerald     #10B981  high score (80+)
  Amber       #F59E0B  medium score (60-79)
  Coral       #F87171  URGENT / low score
  Blue        #3B82F6  COOL / hyperlinks
  Slate       #64748B  subtle text
  Light-gray  #F8FAFC  zebra row banding
"""

from __future__ import annotations

import io
from collections import Counter
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# Title banner — applied to top of every sheet
# ─────────────────────────────────────────────────────────────────────────────


def _apply_title_banner(ws, last_col: str, title: str, subtitle: str):
    """Apply the navy title banner to the top of a worksheet.

    Layout:
      Row 1 (height 56px) — navy bar with bold title
      Row 2 (height 22px) — darker navy with subtitle / metadata
    """
    # Row 1 — main title
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = _font(size=18, bold=True, color=WHITE)
    title_cell.fill = _fill(NAVY)
    title_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        indent=2,
    )
    ws.row_dimensions[1].height = 56

    # Row 2 — subtitle / metadata
    ws.merge_cells(f"A2:{last_col}2")
    sub_cell = ws["A2"]
    sub_cell.value = subtitle
    sub_cell.font = _font(size=10, color="CBD5E1")  # softer slate-300 white
    sub_cell.fill = _fill(NAVY_LIGHT)
    sub_cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[2].height = 22


# ─────────────────────────────────────────────────────────────────────────────
# Color palette (hex without #)
# ─────────────────────────────────────────────────────────────────────────────

NAVY = "0F1D32"
NAVY_LIGHT = "1E3A5F"
WHITE = "FFFFFF"
SLATE = "64748B"
SLATE_DARK = "334155"
LIGHT_BG = "F8FAFC"
ZEBRA = "FAFBFC"
BORDER_LIGHT = "E2E8F0"

GOLD = "D4A853"
SILVER = "C0C0C0"
BRONZE = "B07333"
STONE = "6B665E"

EMERALD = "10B981"
EMERALD_BG = "D1FAE5"
AMBER = "F59E0B"
AMBER_BG = "FEF3C7"
CORAL = "F87171"
CORAL_BG = "FEE2E2"
BLUE = "3B82F6"
BLUE_BG = "DBEAFE"
GRAY_BG = "F1F5F9"


# Tier badge colors — softer pastel backgrounds with dark accent text.
# The full-saturation gold/bronze on a whole cell looked too heavy; pastel
# bg + bold dark accent text reads as a proper chip even when the cell
# fills.
TIER_COLORS = {
    "tier1_ultra_luxury": ("FEF3C7", "92400E"),  # amber-100 bg, amber-800 text
    "tier2_luxury": ("E0E7FF", "3730A3"),  # indigo-100 bg, indigo-800 text
    "tier3_upper_upscale": ("DBEAFE", "1E40AF"),  # blue-100 bg, blue-800 text
    "tier4_upscale": ("F1F5F9", "475569"),  # slate-100 bg, slate-700 text
}

TIER_LABELS = {
    "tier1_ultra_luxury": "Ultra Luxury",
    "tier2_luxury": "Luxury",
    "tier3_upper_upscale": "Upper Upscale",
    "tier4_upscale": "Upscale",
}

# Timeline badge colors — same pastel approach, except URGENT stays bold
# coral because it's intentionally an attention-grabber.
TIMELINE_COLORS = {
    "URGENT": (CORAL, WHITE),  # keep bold — call to action
    "HOT": ("FED7AA", "9A3412"),  # orange-200 / orange-800
    "WARM": ("FEF3C7", "92400E"),  # amber-100 / amber-800
    "COOL": ("DBEAFE", "1E40AF"),  # blue-100 / blue-800
    "TBD": ("F1F5F9", "475569"),
    "EXPIRED": ("E2E8F0", "334155"),
}

# Project type — chip colors + display labels.
# 6 canonical types from app/config/project_type_intelligence.py.
# Renovation (existing property staying open during work) and Reopening
# (was closed, coming back) are different sales contexts so we color
# them differently — not both blue.
PROJECT_TYPE_LABELS = {
    "new_opening": "New Build",
    "renovation": "Renovation",
    "rebrand": "Rebrand",
    "reopening": "Reopening",
    "conversion": "Conversion",
    "ownership_change": "New Owner",
}
PROJECT_TYPE_COLORS = {
    "new_opening": ("D1FAE5", "047857"),  # emerald — clean slate
    "renovation": ("DBEAFE", "1E40AF"),  # blue — existing operation
    "rebrand": ("EDE9FE", "6D28D9"),  # violet — flag change
    "reopening": ("CFFAFE", "0E7490"),  # cyan — restart, distinct from renovation
    "conversion": ("FED7AA", "9A3412"),  # orange — independent → branded
    "ownership_change": ("FEF3C7", "92400E"),  # amber — new owner
}


# ─────────────────────────────────────────────────────────────────────────────
# Style helpers
# ─────────────────────────────────────────────────────────────────────────────

THIN = Side(style="thin", color=BORDER_LIGHT)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _font(size=10, bold=False, color="1F2937", name="Calibri"):
    return Font(name=name, size=size, bold=bold, color=color)


def _fill(color):
    return PatternFill("solid", fgColor=color)


def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left(wrap=False, indent=0):
    return Alignment(
        horizontal="left", vertical="center", wrap_text=wrap, indent=indent
    )


def _right(wrap=False):
    return Alignment(horizontal="right", vertical="center", wrap_text=wrap)


def _score_color(score: int | None) -> tuple[str, str]:
    """Return (bg, fg) for a score badge."""
    if score is None:
        return GRAY_BG, SLATE
    if score >= 80:
        return EMERALD, WHITE
    if score >= 60:
        return AMBER, WHITE
    if score >= 40:
        return STONE, WHITE
    return CORAL, WHITE


def _fmt_currency(val: int | float | None) -> str:
    if val is None or val == 0:
        return "—"
    if val >= 1_000_000:
        return f"${val/1_000_000:.1f}M"
    if val >= 1_000:
        return f"${val/1_000:.0f}K"
    return f"${val:,.0f}"


def _hyperlink_cell(cell, url: str, display: str | None = None):
    """Apply hyperlink formatting to a cell."""
    if not url:
        cell.value = "—"
        cell.font = _font(color=SLATE)
        return
    cell.value = display or url
    cell.hyperlink = url if url.startswith("http") else f"https://{url}"
    cell.font = _font(color=BLUE)


# ─────────────────────────────────────────────────────────────────────────────
# Column specs — defined per export type
# ─────────────────────────────────────────────────────────────────────────────


# Column key, display label, width, getter (called with hotel + primary_contact)
def _g(attr):
    return lambda h, c: getattr(h, attr, None)


def _g_contact(attr):
    return lambda h, c: getattr(c, attr, None) if c else None


def _g_tier_label(h, c):
    t = getattr(h, "brand_tier", None) or ""
    return TIER_LABELS.get(
        t, t.replace("tier", "T").replace("_", " ").title() if t else "—"
    )


def _g_location(h, c):
    parts = [getattr(h, "city", None), getattr(h, "state", None)]
    return ", ".join([p for p in parts if p]) or "—"


def _g_address(h, c):
    parts = [
        getattr(h, "address", None),
        getattr(h, "city", None),
        getattr(h, "state", None),
        getattr(h, "zip_code", None),
    ]
    return ", ".join([p for p in parts if p]) or "—"


def _g_total_contacts(h, c):
    contacts = getattr(h, "_export_all_contacts", None)
    return len(contacts) if contacts else (1 if c else 0)


def _g_client_status(h, c):
    return "Client" if getattr(h, "is_client", False) else "Prospect"


def _g_location(h, c):
    """Quick-read 'City, State' (or 'City, State, Country' for non-US)."""
    city = getattr(h, "city", None) or ""
    state = getattr(h, "state", None) or ""
    country = (getattr(h, "country", None) or "").strip()
    parts = [city, state] if city or state else []
    # Add country only if it's outside USA so domestic rows stay tidy
    if country and country.lower() not in (
        "usa",
        "us",
        "united states",
        "united states of america",
    ):
        parts.append(country)
    return ", ".join([p for p in parts if p]) or "—"


def _g_street_only(h, c):
    """Street address only — empty when not found, full street when known."""
    addr = (getattr(h, "address", None) or "").strip()
    return addr or "—"


def _g_first_source_url(h, c):
    """First URL from source_urls array (or fall back to source_url string).

    Leads can have multiple sources (the source_urls JSON array). We
    expose the first one as a clickable hyperlink — usually the most
    authoritative. Sales who need the full source list go back into
    the app.
    """
    urls = getattr(h, "source_urls", None)
    if urls and isinstance(urls, list):
        first = next((u for u in urls if u), None)
        if first:
            return first
    fallback = getattr(h, "source_url", None)
    if fallback:
        # source_url can be a comma-separated string for legacy rows
        return fallback.split(",")[0].strip() if isinstance(fallback, str) else fallback
    return None


def _g_status(h, c):
    s = (getattr(h, "status", None) or "").strip().lower()
    return {
        "new": "Pipeline",
        "approved": "Approved",
        "rejected": "Rejected",
        "expired": "Expired",
        "deleted": "Deleted",
    }.get(s, s.title() if s else "—")


def _g_date(attr):
    """Return a getter that pulls a datetime field and formats to YYYY-MM-DD."""

    def _getter(h, _c):
        v = getattr(h, attr, None)
        if not v:
            return None
        try:
            return v.strftime("%Y-%m-%d")
        except Exception:
            return str(v)[:10]

    return _getter


# Column key, display label, width, getter (called with hotel + primary_contact)
#
# Order rationale:
#   - score + hotel name first (primary identifiers)
#   - tier/type/project type/timeline (classifiers)
#   - opening + location (when + where)
#   - rooms + revenue (sizing signals)
#   - mgmt/developer/owner (stakeholder rows)
#   - website + address (lookup targets)
#   - contact info (outreach targets)
#   - source URL + status + ID + dates (provenance / cross-ref)
#   - notes (last, free text)

CALL_LIST_COLUMNS_NEW = [
    ("score", "Score", 8, _g("lead_score")),
    ("hotel_name", "Hotel Name", 30, _g("hotel_name")),
    ("brand", "Brand", 18, _g("brand")),
    ("tier", "Tier", 14, _g_tier_label),
    ("type", "Type", 14, _g("hotel_type")),
    ("project_type", "Project Type", 13, _g("project_type")),
    ("timeline", "Timeline", 11, _g("timeline_label")),
    ("opening", "Opening", 14, _g("opening_date")),
    ("location", "Location", 28, _g_location),
    ("zone", "Zone", 20, _g("zone")),
    ("rooms", "Rooms", 8, _g("room_count")),
    ("annual_rev", "Annual Rev", 12, _g("revenue_annual")),
    ("opening_rev", "Opening Order", 13, _g("revenue_opening")),
    ("mgmt", "Mgmt Company", 22, _g("management_company")),
    ("developer", "Developer", 22, _g("developer")),
    ("owner", "Owner", 22, _g("owner")),
    ("website", "Website", 28, _g("hotel_website")),
    ("address", "Street Address", 32, _g_street_only),
    ("c_name", "Contact Name", 20, _g_contact("name")),
    ("c_title", "Contact Title", 22, _g_contact("title")),
    ("c_email", "Contact Email", 26, _g_contact("email")),
    ("c_phone", "Contact Phone", 16, _g_contact("phone")),
    ("c_linkedin", "Contact LinkedIn", 30, _g_contact("linkedin")),
    ("total_contacts", "Total Contacts", 8, _g_total_contacts),
    ("source_url", "Source URL", 32, _g_first_source_url),
    ("status", "Status", 11, _g_status),
    ("lead_id", "ID", 7, _g("id")),
    ("created", "Date Added", 12, _g_date("created_at")),
    ("updated", "Last Updated", 12, _g_date("updated_at")),
    ("notes", "Notes", 40, _g("notes")),
]

CALL_LIST_COLUMNS_EXISTING = [
    ("score", "Score", 8, _g("lead_score")),
    ("hotel_name", "Hotel Name", 30, _g("hotel_name")),
    ("client_status", "Status", 10, _g_client_status),
    ("brand", "Brand", 18, _g("brand")),
    ("tier", "Tier", 14, _g_tier_label),
    ("type", "Type", 14, _g("hotel_type")),
    ("project_type", "Project Type", 13, _g("project_type")),
    ("opening", "Opening (Hist.)", 14, _g("opening_date")),
    ("location", "Location", 28, _g_location),
    ("zone", "Zone", 20, _g("zone")),
    ("rooms", "Rooms", 8, _g("room_count")),
    ("annual_rev", "Annual Rev", 12, _g("revenue_annual")),
    ("mgmt", "Mgmt Company", 22, _g("management_company")),
    ("developer", "Developer", 22, _g("developer")),
    ("owner", "Owner", 22, _g("owner")),
    ("website", "Website", 28, _g("hotel_website")),
    ("address", "Street Address", 32, _g_street_only),
    ("sap_code", "SAP Code", 12, _g("sap_bp_code")),
    ("c_name", "Contact Name", 20, _g_contact("name")),
    ("c_title", "Contact Title", 22, _g_contact("title")),
    ("c_email", "Contact Email", 26, _g_contact("email")),
    ("c_phone", "Contact Phone", 16, _g_contact("phone")),
    ("c_linkedin", "Contact LinkedIn", 30, _g_contact("linkedin")),
    ("total_contacts", "Total Contacts", 8, _g_total_contacts),
    ("source_url", "Source URL", 32, _g_first_source_url),
    ("workflow_status", "Workflow", 11, _g_status),
    ("hotel_id", "ID", 7, _g("id")),
    ("created", "Date Added", 12, _g_date("created_at")),
    ("updated", "Last Updated", 12, _g_date("updated_at")),
    ("notes", "Notes", 40, _g("notes")),
]


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1 — Call List
# ─────────────────────────────────────────────────────────────────────────────


def _build_call_list_sheet(
    wb: Workbook,
    hotels: list,
    primary_contacts: dict[int, Any],
    columns: list[tuple],
    title: str,
    subtitle: str,
):
    ws = wb.active
    ws.title = "Call List"
    ws.sheet_view.showGridLines = False

    n_cols = len(columns)
    last_col = get_column_letter(n_cols)

    # ── Title banner (rows 1-2) ────────────────────────────────────
    _apply_title_banner(ws, last_col, title, subtitle)

    # ── Header row (row 3) ─────────────────────────────────────────
    HEADER_ROW = 3
    for col_idx, (_, label, width, _getter) in enumerate(columns, 1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=label)
        cell.font = _font(size=11, bold=True, color=WHITE)
        cell.fill = _fill(NAVY)
        cell.alignment = _center()
        cell.border = BOX
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[HEADER_ROW].height = 32

    # ── Data rows ──────────────────────────────────────────────────
    DATA_START = HEADER_ROW + 1
    for row_idx, hotel in enumerate(hotels, DATA_START):
        primary = primary_contacts.get(hotel.id)
        is_zebra = (row_idx - DATA_START) % 2 == 1
        zebra_fill = _fill(ZEBRA) if is_zebra else None

        for col_idx, (key, _label, _width, getter) in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = getter(hotel, primary)
            cell.border = BOX

            if key == "score":
                # Colored score badge — rounded look via thick padding
                if value is None:
                    cell.value = "—"
                    cell.font = _font(color=SLATE, bold=True)
                    cell.fill = _fill(GRAY_BG)
                else:
                    bg, fg = _score_color(int(value))
                    cell.value = int(value)
                    cell.font = _font(size=12, bold=True, color=fg)
                    cell.fill = _fill(bg)
                cell.alignment = _center()

            elif key == "tier":
                tier_key = (getattr(hotel, "brand_tier", None) or "").lower()
                bg, fg = TIER_COLORS.get(tier_key, (GRAY_BG, SLATE))
                cell.value = value or "—"
                cell.font = _font(size=10, bold=True, color=fg)
                cell.fill = _fill(bg)
                cell.alignment = _center()

            elif key == "timeline":
                tl = (value or "").upper()
                bg, fg = TIMELINE_COLORS.get(tl, (GRAY_BG, SLATE))
                cell.value = tl or "—"
                cell.font = _font(size=10, bold=True, color=fg)
                cell.fill = _fill(bg)
                cell.alignment = _center()

            elif key == "project_type":
                pt_raw = (getattr(hotel, "project_type", None) or "").lower().strip()
                if pt_raw in PROJECT_TYPE_LABELS:
                    bg, fg = PROJECT_TYPE_COLORS[pt_raw]
                    cell.value = PROJECT_TYPE_LABELS[pt_raw]
                    cell.font = _font(size=10, bold=True, color=fg)
                    cell.fill = _fill(bg)
                else:
                    cell.value = "—"
                    cell.font = _font(color=SLATE)
                    if zebra_fill:
                        cell.fill = zebra_fill
                cell.alignment = _center()

            elif key == "source_url":
                # First source URL only — clickable hyperlink
                if value:
                    display = value
                    if isinstance(display, str):
                        # Trim http/https for cleaner display
                        display = display.replace("https://", "").replace("http://", "")
                        # Truncate visually long URLs but keep the link intact
                        if len(display) > 50:
                            display = display[:47] + "..."
                    _hyperlink_cell(cell, value, display)
                else:
                    cell.value = "—"
                    cell.font = _font(color=SLATE)
                    if zebra_fill:
                        cell.fill = zebra_fill
                cell.alignment = _left()

            elif key in ("status", "workflow_status"):
                # Pipeline / Approved / Rejected / etc. with subtle pastel
                v = (value or "").lower()
                if "pipeline" in v or v == "new":
                    bg, fg = "DBEAFE", "1E40AF"  # blue
                elif "approved" in v:
                    bg, fg = "D1FAE5", "047857"  # emerald
                elif "rejected" in v:
                    bg, fg = "FEE2E2", "B91C1C"  # red
                elif "expired" in v:
                    bg, fg = "F1F5F9", "475569"  # slate
                else:
                    bg, fg = GRAY_BG, SLATE
                cell.value = value or "—"
                cell.font = _font(size=10, bold=True, color=fg)
                cell.fill = _fill(bg)
                cell.alignment = _center()

            elif key in ("lead_id", "hotel_id"):
                cell.value = value if value else "—"
                cell.font = _font(size=10, color=SLATE_DARK, bold=True)
                cell.alignment = _center()
                if zebra_fill:
                    cell.fill = zebra_fill

            elif key in ("created", "updated"):
                cell.value = value if value else "—"
                cell.font = _font(size=10, color=SLATE)
                cell.alignment = _center()
                if zebra_fill:
                    cell.fill = zebra_fill

            elif key == "client_status":
                if value == "Client":
                    cell.fill = _fill(EMERALD_BG)
                    cell.font = _font(size=10, bold=True, color="047857")
                else:
                    cell.fill = _fill(AMBER_BG)
                    cell.font = _font(size=10, bold=True, color="92400E")
                cell.value = value
                cell.alignment = _center()

            elif key in ("annual_rev", "opening_rev"):
                # Store as real number with Excel currency format so pivots
                # and SUM() formulas work in the spreadsheet. Display
                # formatting still produces $1.2M / $450K style readouts
                # via the custom format string.
                if value:
                    cell.value = float(value)
                    cell.number_format = (
                        '[>=1000000]$#,##0.0,,"M";' '[>=1000]$#,##0,"K";' "$#,##0"
                    )
                else:
                    cell.value = "—"
                cell.font = _font(bold=True, color=EMERALD if value else SLATE)
                cell.alignment = _right()
                if zebra_fill:
                    cell.fill = zebra_fill

            elif key in ("rooms", "total_contacts"):
                cell.value = value if value else "—"
                cell.font = _font(color=SLATE_DARK)
                cell.alignment = _center()
                if zebra_fill:
                    cell.fill = zebra_fill

            elif key == "website":
                if value:
                    _hyperlink_cell(
                        cell,
                        value,
                        value.replace("https://", "").replace("http://", ""),
                    )
                else:
                    cell.value = "—"
                    cell.font = _font(color=SLATE)
                cell.alignment = _left()
                if zebra_fill and not value:
                    cell.fill = zebra_fill

            elif key == "c_linkedin":
                if value:
                    _hyperlink_cell(cell, value, "View Profile")
                else:
                    cell.value = "—"
                    cell.font = _font(color=SLATE)
                cell.alignment = _center()
                if zebra_fill and not value:
                    cell.fill = zebra_fill

            elif key == "c_email":
                if value:
                    cell.value = value
                    cell.hyperlink = f"mailto:{value}"
                    cell.font = _font(color=BLUE)
                else:
                    cell.value = "—"
                    cell.font = _font(color=SLATE)
                cell.alignment = _left()
                if zebra_fill and not value:
                    cell.fill = zebra_fill

            elif key in ("hotel_name",):
                cell.value = value or "—"
                cell.font = _font(size=11, bold=True, color="1F2937")
                cell.alignment = _left(wrap=True)
                if zebra_fill:
                    cell.fill = zebra_fill

            elif key == "notes":
                cell.value = value or ""
                cell.font = _font(size=9, color=SLATE_DARK)
                cell.alignment = _left(wrap=True)
                if zebra_fill:
                    cell.fill = zebra_fill

            else:
                cell.value = value if value not in (None, "") else "—"
                cell.font = _font(color=SLATE_DARK if value else SLATE)
                cell.alignment = _left()
                if zebra_fill:
                    cell.fill = zebra_fill

        ws.row_dimensions[row_idx].height = 28

    # ── Freeze panes (header + first column visible while scrolling) ──
    ws.freeze_panes = "C4"

    # ── Auto-filter on header row ──
    if hotels:
        last_data_row = DATA_START + len(hotels) - 1
        ws.auto_filter.ref = f"A{HEADER_ROW}:{last_col}{last_data_row}"


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2 — Summary / Score Distribution (pivot-friendly)
# ─────────────────────────────────────────────────────────────────────────────


def _score_bucket(score: int | None) -> str:
    if score is None:
        return "Unscored"
    if score >= 90:
        return "90-100"
    if score >= 80:
        return "80-89"
    if score >= 70:
        return "70-79"
    if score >= 60:
        return "60-69"
    if score >= 50:
        return "50-59"
    if score >= 40:
        return "40-49"
    return "0-39"


SCORE_BUCKET_ORDER = [
    "90-100",
    "80-89",
    "70-79",
    "60-69",
    "50-59",
    "40-49",
    "0-39",
    "Unscored",
]


def _section_title(ws, row: int, last_col: str, text: str):
    ws.merge_cells(f"A{row}:{last_col}{row}")
    cell = ws[f"A{row}"]
    cell.value = text
    cell.font = _font(size=13, bold=True, color=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28


def _table_header(ws, row: int, headers: list[str]):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = _font(bold=True, color=WHITE, size=10)
        cell.fill = _fill(NAVY)
        cell.alignment = _center()
        cell.border = BOX
    ws.row_dimensions[row].height = 24


def _table_row(ws, row: int, values: list, n_cols: int, accent_first=False):
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        if col_idx <= len(values):
            cell.value = values[col_idx - 1]
        cell.border = BOX
        cell.font = _font(
            size=10, color=SLATE_DARK, bold=(accent_first and col_idx == 1)
        )
        cell.alignment = _center() if col_idx > 1 else _left(indent=1)


def _build_summary_sheet(
    wb: Workbook,
    hotels: list,
    primary_contacts: dict[int, Any],
    *,
    include_timeline: bool,
    label: str,
):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18

    # ── Title banner ──
    _apply_title_banner(
        ws,
        "D",
        f"{label} — Summary & Score Distribution",
        f"Total: {len(hotels)} hotel(s)",
    )

    row = 4
    total = len(hotels) or 1

    # ── Score Distribution ───────────────────────────────────────────
    _section_title(ws, row, "D", "Score Distribution")
    row += 1
    _table_header(ws, row, ["Score Range", "Count", "% of Total", "Cumulative"])
    row += 1
    score_counts = Counter(
        _score_bucket(getattr(h, "lead_score", None)) for h in hotels
    )
    cumulative = 0
    for bucket in SCORE_BUCKET_ORDER:
        n = score_counts.get(bucket, 0)
        if n == 0:
            continue
        cumulative += n
        pct = f"{100 * n / total:.1f}%"
        cum_pct = f"{cumulative} ({100*cumulative/total:.0f}%)"
        _table_row(ws, row, [bucket, n, pct, cum_pct], 4, accent_first=True)
        # Colored score-range cell
        bg, fg = _score_color(
            {
                "90-100": 95,
                "80-89": 85,
                "70-79": 75,
                "60-69": 65,
                "50-59": 55,
                "40-49": 45,
                "0-39": 30,
                "Unscored": None,
            }.get(bucket, None)
        )
        ws.cell(row=row, column=1).fill = _fill(bg)
        ws.cell(row=row, column=1).font = _font(size=10, bold=True, color=fg)
        row += 1
    row += 2

    # ── By Tier ──────────────────────────────────────────────────────
    _section_title(ws, row, "D", "By Brand Tier")
    row += 1
    _table_header(ws, row, ["Tier", "Count", "Avg Score", "Total Annual Rev"])
    row += 1
    tier_buckets: dict[str, list] = {}
    for h in hotels:
        t = getattr(h, "brand_tier", None) or "Unknown"
        tier_buckets.setdefault(t, []).append(h)
    tier_order = [
        "tier1_ultra_luxury",
        "tier2_luxury",
        "tier3_upper_upscale",
        "tier4_upscale",
        "Unknown",
    ]
    for t in tier_order + [k for k in tier_buckets if k not in tier_order]:
        bucket = tier_buckets.get(t, [])
        if not bucket:
            continue
        avg_score = round(
            sum(getattr(h, "lead_score", 0) or 0 for h in bucket) / len(bucket), 1
        )
        total_rev = sum(getattr(h, "revenue_annual", 0) or 0 for h in bucket)
        label_display = TIER_LABELS.get(t, t.replace("_", " ").title())
        _table_row(
            ws,
            row,
            [label_display, len(bucket), avg_score, _fmt_currency(total_rev)],
            4,
            accent_first=True,
        )
        # Tier badge color on first cell
        bg, fg = TIER_COLORS.get(t, (GRAY_BG, SLATE))
        ws.cell(row=row, column=1).fill = _fill(bg)
        ws.cell(row=row, column=1).font = _font(size=10, bold=True, color=fg)
        row += 1
    row += 2

    # ── By Timeline (New Hotels only) ────────────────────────────────
    if include_timeline:
        _section_title(ws, row, "D", "By Timeline")
        row += 1
        _table_header(ws, row, ["Timeline", "Count", "Avg Score", "% of Total"])
        row += 1
        tl_buckets: dict[str, list] = {}
        for h in hotels:
            tl = (getattr(h, "timeline_label", None) or "TBD").upper()
            tl_buckets.setdefault(tl, []).append(h)
        for tl in ["URGENT", "HOT", "WARM", "COOL", "TBD", "EXPIRED"]:
            bucket = tl_buckets.get(tl, [])
            if not bucket:
                continue
            avg_score = round(
                sum(getattr(h, "lead_score", 0) or 0 for h in bucket) / len(bucket), 1
            )
            pct = f"{100 * len(bucket) / total:.1f}%"
            _table_row(ws, row, [tl, len(bucket), avg_score, pct], 4, accent_first=True)
            bg, fg = TIMELINE_COLORS.get(tl, (GRAY_BG, SLATE))
            ws.cell(row=row, column=1).fill = _fill(bg)
            ws.cell(row=row, column=1).font = _font(size=10, bold=True, color=fg)
            row += 1
        row += 2

    # ── By State (top 15) ────────────────────────────────────────────
    _section_title(ws, row, "D", "Top 15 States")
    row += 1
    _table_header(ws, row, ["State", "Count", "Avg Score", "Avg Rooms"])
    row += 1
    state_counts: dict[str, list] = {}
    for h in hotels:
        s = getattr(h, "state", None) or "—"
        state_counts.setdefault(s, []).append(h)
    top_states = sorted(state_counts.items(), key=lambda x: -len(x[1]))[:15]
    for state, bucket in top_states:
        avg_score = round(
            sum(getattr(h, "lead_score", 0) or 0 for h in bucket) / len(bucket), 1
        )
        rooms = [getattr(h, "room_count", 0) or 0 for h in bucket]
        avg_rooms = round(sum(rooms) / len(rooms)) if rooms else 0
        _table_row(
            ws,
            row,
            [state, len(bucket), avg_score, avg_rooms or "—"],
            4,
            accent_first=True,
        )
        row += 1
    row += 2

    # ── By Zone (≥3 leads) ───────────────────────────────────────────
    zone_counts: dict[str, list] = {}
    for h in hotels:
        z = getattr(h, "zone", None) or "—"
        zone_counts.setdefault(z, []).append(h)
    zone_table = sorted(
        [(z, b) for z, b in zone_counts.items() if len(b) >= 3],
        key=lambda x: -len(x[1]),
    )
    if zone_table:
        _section_title(ws, row, "D", "By Zone (3+ leads)")
        row += 1
        _table_header(ws, row, ["Zone", "Count", "Avg Score", "% of Total"])
        row += 1
        for zone, bucket in zone_table:
            avg_score = round(
                sum(getattr(h, "lead_score", 0) or 0 for h in bucket) / len(bucket), 1
            )
            pct = f"{100 * len(bucket) / total:.1f}%"
            _table_row(
                ws, row, [zone, len(bucket), avg_score, pct], 4, accent_first=True
            )
            row += 1
        row += 2

    # ── Contact Coverage ─────────────────────────────────────────────
    _section_title(ws, row, "D", "Contact Coverage")
    row += 1
    _table_header(ws, row, ["Coverage Type", "Count", "% of Total", ""])
    row += 1

    has_email = sum(
        1
        for h in hotels
        if (
            primary_contacts.get(h.id)
            and getattr(primary_contacts.get(h.id), "email", None)
        )
    )
    has_phone = sum(
        1
        for h in hotels
        if (
            primary_contacts.get(h.id)
            and getattr(primary_contacts.get(h.id), "phone", None)
        )
    )
    has_linkedin = sum(
        1
        for h in hotels
        if (
            primary_contacts.get(h.id)
            and getattr(primary_contacts.get(h.id), "linkedin", None)
        )
    )
    has_any = sum(1 for h in hotels if primary_contacts.get(h.id))
    no_contact = total - has_any

    for label_text, count, color in [
        ("Has primary email", has_email, EMERALD_BG),
        ("Has primary phone", has_phone, EMERALD_BG),
        ("Has primary LinkedIn", has_linkedin, EMERALD_BG),
        ("Has at least one contact", has_any, BLUE_BG),
        ("No contact yet", no_contact, CORAL_BG),
    ]:
        pct = f"{100 * count / total:.1f}%" if total else "0%"
        _table_row(ws, row, [label_text, count, pct, ""], 4, accent_first=True)
        ws.cell(row=row, column=1).fill = _fill(color)
        row += 1


# ─────────────────────────────────────────────────────────────────────────────
# Sheet — All Contacts (one row per contact)
# ─────────────────────────────────────────────────────────────────────────────

# Priority badge colors for P1-P4 (matches the in-app priority chips)
PRIORITY_COLORS = {
    "P1": ("D1FAE5", "047857"),  # Call first — emerald pastel
    "P2": ("DBEAFE", "1E40AF"),  # Strong fit — blue pastel
    "P3": ("FEF3C7", "92400E"),  # Useful — amber pastel
    "P4": ("F1F5F9", "475569"),  # Escalation only — slate pastel
}

# Scope colors (matches the chip styling in the contact panel)
SCOPE_COLORS = {
    "hotel_specific": (EMERALD_BG, "047857"),
    "chain_area": (AMBER_BG, "92400E"),
    "management_corporate": (BLUE_BG, "1E40AF"),
    "chain_corporate": ("F1F5F9", SLATE),
    "owner": ("F3E8FF", "6B21A8"),
}


def _build_all_contacts_sheet(wb: Workbook, hotels: list, *, kind: str, tab_label: str):
    """Build the All Contacts sheet — one row per contact across every hotel.

    Each contact row carries hotel context (name, brand, tier, score) so a
    sales rep can sort/filter freely without losing track of which hotel a
    person belongs to. P1-P4 priority badges and scope chips match the UI.
    """
    ws = wb.create_sheet("All Contacts")
    ws.sheet_view.showGridLines = False

    HEADERS = [
        ("Hotel Name", 28),
        ("Hotel Score", 10),
        ("Brand", 18),
        ("Tier", 14),
        ("City", 16),
        ("State", 8),
        ("Priority", 9),
        ("Contact Name", 20),
        ("Title", 24),
        ("Organization", 22),
        ("Scope", 18),
        ("Score", 8),
        ("Email", 26),
        ("Phone", 16),
        ("LinkedIn", 28),
        ("Found Via", 14),
    ]
    n_cols = len(HEADERS)
    last_col = get_column_letter(n_cols)

    # Count contacts for subtitle
    total_contacts = sum(
        len(getattr(h, "_export_all_contacts", []) or []) for h in hotels
    )

    # ── Title banner ──
    label = "New Hotels" if kind == "new" else "Existing Hotels"
    _apply_title_banner(
        ws,
        last_col,
        f"JA Uniforms — {label} Contacts ({tab_label})",
        f"{total_contacts} contact(s) across {len(hotels)} hotel(s) · "
        f"sort/filter freely for batch outreach",
    )

    # ── Header row (row 3) ──
    HEADER_ROW = 3
    for col_idx, (label_text, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=label_text)
        cell.font = _font(size=11, bold=True, color=WHITE)
        cell.fill = _fill(NAVY)
        cell.alignment = _center()
        cell.border = BOX
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[HEADER_ROW].height = 32

    # ── Data rows ──
    DATA_START = HEADER_ROW + 1
    row = DATA_START
    for hotel in hotels:
        contacts = getattr(hotel, "_export_all_contacts", None) or []
        if not contacts:
            continue

        h_name = (
            getattr(hotel, "hotel_name", None) or getattr(hotel, "name", None) or "—"
        )
        h_score = getattr(hotel, "lead_score", None)
        h_brand = getattr(hotel, "brand", None) or "—"
        h_tier_key = (getattr(hotel, "brand_tier", None) or "").lower()
        h_tier_label = TIER_LABELS.get(h_tier_key, "—")
        h_city = getattr(hotel, "city", None) or "—"
        h_state = getattr(hotel, "state", None) or "—"

        for c in contacts:
            is_zebra = (row - DATA_START) % 2 == 1
            zebra_fill = _fill(ZEBRA) if is_zebra else None

            # Hotel context columns
            cell = ws.cell(row=row, column=1, value=h_name)
            cell.font = _font(size=10, bold=True, color="1F2937")
            cell.alignment = _left(wrap=True)
            cell.border = BOX
            if zebra_fill:
                cell.fill = zebra_fill

            # Hotel score badge
            sc = ws.cell(row=row, column=2)
            if h_score is None:
                sc.value = "—"
                sc.font = _font(color=SLATE)
                sc.fill = _fill(GRAY_BG)
            else:
                bg, fg = _score_color(int(h_score))
                sc.value = int(h_score)
                sc.font = _font(size=11, bold=True, color=fg)
                sc.fill = _fill(bg)
            sc.alignment = _center()
            sc.border = BOX

            # Brand
            cell = ws.cell(row=row, column=3, value=h_brand)
            cell.font = _font(color=SLATE_DARK)
            cell.alignment = _left()
            cell.border = BOX
            if zebra_fill:
                cell.fill = zebra_fill

            # Tier chip
            tier_cell = ws.cell(row=row, column=4, value=h_tier_label)
            bg, fg = TIER_COLORS.get(h_tier_key, (GRAY_BG, SLATE))
            tier_cell.font = _font(size=10, bold=True, color=fg)
            tier_cell.fill = _fill(bg)
            tier_cell.alignment = _center()
            tier_cell.border = BOX

            # City / State
            for col_idx, val in [(5, h_city), (6, h_state)]:
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = _font(color=SLATE_DARK)
                cell.alignment = _left() if col_idx == 5 else _center()
                cell.border = BOX
                if zebra_fill:
                    cell.fill = zebra_fill

            # Priority badge (P1-P4) — computed live since it's not a column
            try:
                priority, _reason = c._compute_priority()
            except Exception:
                priority = "—"
            p_cell = ws.cell(row=row, column=7, value=priority)
            if priority in PRIORITY_COLORS:
                bg, fg = PRIORITY_COLORS[priority]
                p_cell.font = _font(size=10, bold=True, color=fg)
                p_cell.fill = _fill(bg)
            else:
                p_cell.font = _font(color=SLATE)
                p_cell.fill = _fill(GRAY_BG)
            p_cell.alignment = _center()
            p_cell.border = BOX

            # Contact name (bold, primary contacts get a star)
            name_val = getattr(c, "name", None) or "—"
            is_primary = bool(getattr(c, "is_primary", False))
            display_name = f"★ {name_val}" if is_primary else name_val
            n_cell = ws.cell(row=row, column=8, value=display_name)
            n_cell.font = _font(size=10, bold=True, color="1F2937")
            n_cell.alignment = _left()
            n_cell.border = BOX
            if zebra_fill:
                n_cell.fill = zebra_fill

            # Title
            title_val = getattr(c, "title", None) or "—"
            t_cell = ws.cell(row=row, column=9, value=title_val)
            t_cell.font = _font(color=SLATE_DARK)
            t_cell.alignment = _left(wrap=True)
            t_cell.border = BOX
            if zebra_fill:
                t_cell.fill = zebra_fill

            # Organization
            org = getattr(c, "organization", None) or "—"
            o_cell = ws.cell(row=row, column=10, value=org)
            o_cell.font = _font(color=SLATE_DARK)
            o_cell.alignment = _left()
            o_cell.border = BOX
            if zebra_fill:
                o_cell.fill = zebra_fill

            # Scope chip
            scope = (getattr(c, "scope", None) or "").lower()
            scope_label = scope.replace("_", " ").title() if scope else "—"
            s_cell = ws.cell(row=row, column=11, value=scope_label)
            if scope in SCOPE_COLORS:
                bg, fg = SCOPE_COLORS[scope]
                s_cell.font = _font(size=10, bold=True, color=fg)
                s_cell.fill = _fill(bg)
            else:
                s_cell.font = _font(color=SLATE)
                if zebra_fill:
                    s_cell.fill = zebra_fill
            s_cell.alignment = _center()
            s_cell.border = BOX

            # Contact score
            c_score = getattr(c, "score", None)
            cs_cell = ws.cell(row=row, column=12, value=c_score if c_score else "—")
            cs_cell.font = _font(
                bold=True,
                color=EMERALD
                if (c_score or 0) >= 60
                else AMBER
                if (c_score or 0) >= 30
                else SLATE,
            )
            cs_cell.alignment = _center()
            cs_cell.border = BOX
            if zebra_fill:
                cs_cell.fill = zebra_fill

            # Email (mailto: hyperlink)
            email = getattr(c, "email", None)
            e_cell = ws.cell(row=row, column=13)
            if email:
                e_cell.value = email
                e_cell.hyperlink = f"mailto:{email}"
                e_cell.font = _font(color=BLUE)
            else:
                e_cell.value = "—"
                e_cell.font = _font(color=SLATE)
                if zebra_fill:
                    e_cell.fill = zebra_fill
            e_cell.alignment = _left()
            e_cell.border = BOX

            # Phone
            phone = getattr(c, "phone", None) or "—"
            ph_cell = ws.cell(row=row, column=14, value=phone)
            ph_cell.font = _font(color=SLATE_DARK if phone != "—" else SLATE)
            ph_cell.alignment = _left()
            ph_cell.border = BOX
            if zebra_fill:
                ph_cell.fill = zebra_fill

            # LinkedIn
            linkedin = getattr(c, "linkedin", None)
            li_cell = ws.cell(row=row, column=15)
            if linkedin:
                _hyperlink_cell(li_cell, linkedin, "View Profile")
            else:
                li_cell.value = "—"
                li_cell.font = _font(color=SLATE)
                if zebra_fill:
                    li_cell.fill = zebra_fill
            li_cell.alignment = _center()
            li_cell.border = BOX

            # Found Via (wiza_valid / serper / manual / etc)
            fv = getattr(c, "found_via", None) or "—"
            fv_label = fv.replace("_", " ").title() if fv != "—" else fv
            fv_cell = ws.cell(row=row, column=16, value=fv_label)
            fv_cell.font = _font(size=9, color=SLATE)
            fv_cell.alignment = _center()
            fv_cell.border = BOX
            if zebra_fill:
                fv_cell.fill = zebra_fill

            ws.row_dimensions[row].height = 26
            row += 1

    # ── Freeze + autofilter ──
    ws.freeze_panes = "C4"
    if row > DATA_START:
        ws.auto_filter.ref = f"A{HEADER_ROW}:{last_col}{row - 1}"


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────


def build_workbook(
    hotels: list,
    primary_contacts: dict[int, Any],
    *,
    kind: str,
    tab_label: str = "Pipeline",
) -> bytes:
    """Build a polished 3-sheet Excel workbook.

    Args:
        hotels: list of PotentialLead OR ExistingHotel ORM rows
        primary_contacts: {hotel.id: highest-scoring LeadContact} mapping
        kind: "new" or "existing"
        tab_label: which tab the user was viewing (Pipeline/Approved/Rejected),
                   used in the title banner and filename

    Sheets:
        1. Call List       — one row per hotel + primary contact
        2. All Contacts    — one row per contact, hotel info repeated.
                             Sales sorts/filters this for batch outreach.
        3. Summary         — pivot-friendly aggregations (boss view)

    Returns:
        bytes — ready to stream as
        application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
    """
    wb = Workbook()

    if kind == "new":
        columns = CALL_LIST_COLUMNS_NEW
        sheet_title = f"JA Uniforms — New Hotels Pipeline ({tab_label})"
        summary_label = "New Hotels"
    else:
        columns = CALL_LIST_COLUMNS_EXISTING
        sheet_title = f"JA Uniforms — Existing Hotels ({tab_label})"
        summary_label = "Existing Hotels"

    from app.services.utils import local_now

    subtitle = (
        f"{len(hotels)} hotel(s) · "
        f"Exported {local_now().strftime('%B %d, %Y at %I:%M %p')}"
    )

    _build_call_list_sheet(
        wb,
        hotels,
        primary_contacts,
        columns,
        sheet_title,
        subtitle,
    )
    _build_all_contacts_sheet(wb, hotels, kind=kind, tab_label=tab_label)
    _build_summary_sheet(
        wb,
        hotels,
        primary_contacts,
        include_timeline=(kind == "new"),
        label=summary_label,
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
