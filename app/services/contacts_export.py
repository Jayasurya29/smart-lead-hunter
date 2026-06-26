"""contacts_export.py — build the contacts directory as a polished .xlsx.

Read-only. Excludes trash (manual_category OR contact_category = 'junk') and
operational shared inboxes. Two tabs:
  • Overview  — totals, category/priority/vertical breakdowns, completeness gaps
  • Contacts  — one styled row per contact, buyer-first then priority then org,
                colour-coded category/priority, hyperlinked email + LinkedIn,
                zebra banding, frozen header + autofilter.
"""

from __future__ import annotations

from io import BytesIO

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

# ── palette (matches the leads export house style) ──
NAVY = "2E4A6E"
NAVY_LIGHT = "3E5A80"
WHITE = "FFFFFF"
SLATE = "64748B"
INK = "1F2937"
ZEBRA = "F4F7FB"  # very light navy-tint stripe
BORDER = "E2E8F0"
EMERALD = "1A7A55"
AMBER = "C49A3C"
CORAL = "E85D4A"
BLUE = "3E638C"
GRAY = "9CA3AF"

_CAT_COLOR = {
    "buyer": EMERALD,
    "seller": AMBER,
    "competitor": CORAL,
    "personal": BLUE,
    "prospect": "5B7A9E",
}
_PRIO_COLOR = {"P1": EMERALD, "P2": "5E9C6E", "P3": AMBER, "P4": GRAY}

_WHERE = (
    "WHERE COALESCE(c.manual_category, c.contact_category) IS DISTINCT FROM 'junk' "
    "AND COALESCE(c.manual_category, c.contact_category) IS DISTINCT FROM 'operational' "
    "AND COALESCE(c.is_shared_mailbox, false) = false"
)

_SQL = """
    SELECT c.id, c.display_name, c.email,
           c.organization, c.parent_company, c.management_company,
           COALESCE(NULLIF(c.title,''), c.inferred_role) AS role,
           c.seniority, c.department,
           COALESCE(c.manual_category, c.contact_category) AS category,
           c.procurement_priority,
           r.vertical AS dict_vertical,
           c.is_decision_maker, c.brand_tier, c.gpo,
           c.linkedin_url, c.phone,
           c.interaction_count, c.relevance_verdict,
           c.enrichment_source, to_char(c.enriched_at,'YYYY-MM-DD') AS enriched_at,
           c.org_source, c.approval_status,
           array_to_string(c.source_mailboxes, '; ') AS mailboxes,
           to_char(c.first_seen,'YYYY-MM-DD') AS first_seen,
           to_char(c.last_seen,'YYYY-MM-DD') AS last_seen
    FROM contacts c
    LEFT JOIN contact_roles r
      ON r.role_normalized = lower(regexp_replace(
             regexp_replace(COALESCE(NULLIF(c.title,''), c.inferred_role, ''),
             '&', ' and ', 'g'), '[^a-zA-Z0-9 ]+', ' ', 'g'))
    {where}
    ORDER BY
      CASE COALESCE(c.manual_category, c.contact_category)
        WHEN 'buyer' THEN 0 WHEN 'seller' THEN 2 WHEN 'competitor' THEN 3
        WHEN 'personal' THEN 4 ELSE 1 END,
      CASE c.procurement_priority WHEN 'P1' THEN 0 WHEN 'P2' THEN 1
        WHEN 'P3' THEN 2 WHEN 'P4' THEN 3 ELSE 4 END,
      c.organization, c.id
"""

# (attr, header, width, align)  align: l/c
_COLS = [
    ("display_name", "Name", 26, "l"),
    ("email", "Email", 32, "l"),
    ("organization", "Organization", 28, "l"),
    ("role", "Role", 26, "l"),
    ("seniority", "Seniority", 11, "c"),
    ("category", "Category", 12, "c"),
    ("procurement_priority", "Priority", 9, "c"),
    ("dict_vertical", "Vertical", 13, "c"),
    ("is_decision_maker", "DM", 5, "c"),
    ("brand_tier", "Brand tier", 16, "l"),
    ("gpo", "GPO", 10, "c"),
    ("linkedin_url", "LinkedIn", 16, "c"),
    ("phone", "Phone", 16, "l"),
    ("interaction_count", "Emails", 7, "c"),
    ("relevance_verdict", "Relevance", 11, "c"),
    ("enrichment_source", "Enriched via", 12, "c"),
    ("enriched_at", "Enriched", 11, "c"),
    ("parent_company", "Parent co", 22, "l"),
    ("management_company", "Mgmt co", 22, "l"),
    ("department", "Dept", 16, "l"),
    ("org_source", "Org source", 13, "c"),
    ("approval_status", "Status", 14, "c"),
    ("mailboxes", "Source mailboxes", 30, "l"),
    ("first_seen", "First seen", 11, "c"),
    ("last_seen", "Last seen", 11, "c"),
]


async def build_contacts_xlsx(session: AsyncSession) -> bytes:
    """Return the contacts directory as polished .xlsx bytes (trash excluded)."""
    from datetime import datetime

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    rows = (await session.execute(text(_SQL.format(where=_WHERE)))).mappings().all()

    def F(size=10, bold=False, color=INK, name="Calibri"):
        return Font(name=name, size=size, bold=bold, color=color)

    def Fill(c):
        return PatternFill("solid", fgColor=c)

    left = Alignment(horizontal="left", vertical="center", indent=1)
    center = Alignment(horizontal="center", vertical="center")

    # [perf] PRE-BUILD every style object ONCE and reuse. Creating fresh Font/
    # Fill/Alignment per cell over ~800k cells took ~70s; sharing instances drops
    # it to a few seconds. openpyxl is fast when styles are shared.
    _zebra_fill = Fill(ZEBRA)
    _font_ink = F(size=10, color=INK)
    _font_blue = F(size=10, color=BLUE)
    _font_dm = F(size=11, bold=True, color=EMERALD)
    _font_cat = {k: F(size=10, bold=True, color=v) for k, v in _CAT_COLOR.items()}
    _font_cat_default = F(size=10, bold=True, color=SLATE)
    _font_prio = {k: F(size=10, bold=True, color=v) for k, v in _PRIO_COLOR.items()}
    _font_prio_default = F(size=10, bold=True, color=GRAY)

    wb = Workbook()

    # ══ Contacts sheet ══
    ws = wb.active
    ws.title = "Contacts"
    ncols = len(_COLS)
    last_col = get_column_letter(ncols)

    # title banner (rows 1-2)
    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value = "JA Uniforms  —  Contact Directory"
    t.font = F(size=18, bold=True, color=WHITE)
    t.fill = Fill(NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[1].height = 42

    ws.merge_cells(f"A2:{last_col}2")
    s = ws["A2"]
    s.value = (
        f"{len(rows):,} contacts  ·  trash excluded  ·  " f"exported {datetime.now():%B %d, %Y}"
    )
    s.font = F(size=10, color="CBD5E1")
    s.fill = Fill(NAVY_LIGHT)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[2].height = 20

    # header row (row 3)
    hr = 3
    for ci, (_a, label, width, _al) in enumerate(_COLS, start=1):
        c = ws.cell(row=hr, column=ci, value=label)
        c.font = F(size=10, bold=True, color=WHITE)
        c.fill = Fill(NAVY)
        c.alignment = center
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[hr].height = 22

    # data rows
    for i, row in enumerate(rows):
        r = hr + 1 + i
        zebra = i % 2 == 1
        for ci, (attr, _label, _w, al) in enumerate(_COLS, start=1):
            v = row.get(attr)
            cell = ws.cell(row=r, column=ci)
            cell.alignment = left if al == "l" else center
            if zebra:
                cell.fill = _zebra_fill

            if attr == "is_decision_maker":
                cell.value = "✓" if v else ""
                cell.font = _font_dm
            elif attr == "category" and v:
                cell.value = str(v).title()
                cell.font = _font_cat.get(v, _font_cat_default)
            elif attr == "procurement_priority" and v:
                cell.value = v
                cell.font = _font_prio.get(v, _font_prio_default)
            elif attr == "email" and v:
                cell.value = v
                cell.hyperlink = f"mailto:{v}"
                cell.font = _font_blue
            elif attr == "linkedin_url" and v:
                cell.value = "LinkedIn"
                cell.hyperlink = v
                cell.font = _font_blue
            else:
                cell.value = v
                cell.font = _font_ink

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A{hr}:{last_col}{hr + len(rows)}"

    # ══ Overview sheet ══
    ov = wb.create_sheet("Overview", 0)  # first tab
    ov.column_dimensions["A"].width = 26
    ov.column_dimensions["B"].width = 14
    ov.column_dimensions["C"].width = 14
    ov.merge_cells("A1:C1")
    h = ov["A1"]
    h.value = "Contact Directory — Overview"
    h.font = F(size=16, bold=True, color=WHITE)
    h.fill = Fill(NAVY)
    h.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ov.row_dimensions[1].height = 36

    def section(rownum, title):
        ov.merge_cells(f"A{rownum}:C{rownum}")
        c = ov.cell(rownum, 1, title)
        c.font = F(size=11, bold=True, color=NAVY)
        c.fill = Fill("EEF2F7")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ov.row_dimensions[rownum].height = 20
        return rownum + 1

    def kv(rownum, k, v, color=INK, bold=False):
        ov.cell(rownum, 1, k).font = F(size=10, color=SLATE)
        c = ov.cell(rownum, 2, v)
        c.font = F(size=10, bold=bold, color=color)
        c.alignment = center
        return rownum + 1

    total = len(rows)
    from collections import Counter

    cats = Counter(r.get("category") or "uncategorized" for r in rows)
    prios = Counter(r.get("procurement_priority") or "—" for r in rows)
    verts = Counter(r.get("dict_vertical") or "other" for r in rows)
    dms = sum(1 for r in rows if r.get("is_decision_maker"))

    def miss(pred):
        n = sum(1 for r in rows if pred(r))
        pct = round(100 * n / total) if total else 0
        return f"{n:,}  ({pct}%)"

    rn = 3
    rn = section(rn, "Totals")
    rn = kv(rn, "Total contacts", f"{total:,}", color=NAVY, bold=True)
    rn = kv(rn, "Decision-makers", f"{dms:,}", color=EMERALD, bold=True)
    rn += 1

    rn = section(rn, "By category")
    for cat in ["buyer", "seller", "competitor", "personal", "prospect", "uncategorized"]:
        if cats.get(cat):
            rn = kv(
                rn,
                cat.title(),
                f"{cats[cat]:,}",
                color=_CAT_COLOR.get(cat, SLATE),
                bold=(cat == "buyer"),
            )
    rn += 1

    rn = section(rn, "By priority")
    for p in ["P1", "P2", "P3", "P4", "P_unknown", "—"]:
        if prios.get(p):
            rn = kv(rn, p, f"{prios[p]:,}", color=_PRIO_COLOR.get(p, GRAY))
    rn += 1

    rn = section(rn, "By vertical")
    for v, n in verts.most_common():
        rn = kv(rn, str(v).replace("_", " ").title(), f"{n:,}")
    rn += 1

    rn = section(rn, "Data completeness (missing)")
    rn = kv(rn, "Missing email", miss(lambda r: not r.get("email")))
    rn = kv(rn, "Missing role", miss(lambda r: not r.get("role")))
    rn = kv(rn, "Missing LinkedIn", miss(lambda r: not r.get("linkedin_url")))
    rn = kv(rn, "Missing phone", miss(lambda r: not r.get("phone")))
    rn = kv(rn, "Missing organization", miss(lambda r: not r.get("organization")))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
