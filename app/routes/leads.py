"""Lead CRUD endpoints — list, detail, create, update."""

import logging
from datetime import timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.database import get_db
from app.models import PotentialLead
from app.schemas import (
    LeadCreate,
    LeadUpdate,
    LeadResponse,
    LeadListResponse,
)
from app.services.lead_factory import save_lead_to_db
from app.services.utils import local_now
from app.shared import (
    apply_lead_filters,
    paginate_leads,
    lead_list_response,
    escape_like,
)

logger = logging.getLogger(__name__)

router = APIRouter()


# -----------------------------------------------------------------------------
# Lead List (full filtering for React frontend)
# -----------------------------------------------------------------------------


@router.get("/leads", response_model=LeadListResponse, tags=["Leads"])
async def list_leads(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=500),
    status: Optional[str] = None,
    min_score: Optional[int] = None,
    state: Optional[str] = None,
    location_type: Optional[str] = None,
    brand_tier: Optional[str] = None,
    search: Optional[str] = None,
    timeline: Optional[str] = None,
    year: Optional[str] = None,
    added: Optional[str] = None,
    sort: Optional[str] = "score_desc",
    location: Optional[str] = None,
    has_coords: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    """List leads with full filtering, pagination, and sorting."""
    query = select(PotentialLead)
    count_query = select(func.count(PotentialLead.id))

    query, count_query = apply_lead_filters(
        query,
        count_query,
        status=status,
        min_score=min_score,
        state=state,
        location_type=location_type,
        brand_tier=brand_tier,
        search=search,
    )

    now = local_now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)

    # Timeline filter (uses precomputed column)
    if timeline and timeline in (
        "hot",
        "urgent",
        "warm",
        "cool",
        "late",
        "expired",
        "tbd",
    ):
        label = timeline.upper()
        query = query.where(PotentialLead.timeline_label == label)
        count_query = count_query.where(PotentialLead.timeline_label == label)

    # Year filter
    if year:
        safe_year = escape_like(year)
        query = query.where(PotentialLead.opening_date.ilike(f"%{safe_year}%"))
        count_query = count_query.where(
            PotentialLead.opening_date.ilike(f"%{safe_year}%")
        )

    # has_coords filter — used by map to get only geocoded leads
    if has_coords == "true":
        query = query.where(
            PotentialLead.latitude.isnot(None),
            PotentialLead.longitude.isnot(None),
        )
        count_query = count_query.where(
            PotentialLead.latitude.isnot(None),
            PotentialLead.longitude.isnot(None),
        )

    # Added date filter
    if added:
        if added == "today":
            cutoff = today_start
        elif added == "this_week":
            cutoff = today_start - timedelta(days=now.weekday())
        elif added == "last_7":
            cutoff = today_start - timedelta(days=7)
        elif added == "last_30":
            cutoff = today_start - timedelta(days=30)
        else:
            cutoff = None
        if cutoff:
            query = query.where(PotentialLead.created_at >= cutoff)
            count_query = count_query.where(PotentialLead.created_at >= cutoff)

    # Location filter
    if location:
        from app.config.locations import (
            SOUTH_FLORIDA_CITIES,
            CARIBBEAN_COUNTRIES,
            SOUTHEAST_STATES,
            MOUNTAIN_STATES,
            NORTHEAST_STATES,
            DC_NAMES,
            MIDWEST_STATES,
            PACIFIC_NW_STATES,
            LAS_VEGAS_CITIES,
            NEW_ORLEANS_CITIES,
            HAWAII_CITIES,
        )

        loc_filter = None
        if location == "south_florida":
            loc_filter = func.lower(PotentialLead.city).in_(SOUTH_FLORIDA_CITIES)
        elif location == "rest_florida":
            loc_filter = (func.lower(PotentialLead.state) == "florida") & ~func.lower(
                PotentialLead.city
            ).in_(SOUTH_FLORIDA_CITIES)
        elif location == "caribbean":
            loc_filter = func.lower(PotentialLead.country).in_(CARIBBEAN_COUNTRIES)
        elif location == "california":
            loc_filter = func.lower(PotentialLead.state) == "california"
        elif location == "new_york":
            loc_filter = func.lower(PotentialLead.state) == "new york"
        elif location == "texas":
            loc_filter = func.lower(PotentialLead.state) == "texas"
        elif location == "southeast":
            loc_filter = func.lower(PotentialLead.state).in_(SOUTHEAST_STATES)
        elif location == "mountain":
            loc_filter = func.lower(PotentialLead.state).in_(MOUNTAIN_STATES)
        elif location == "northeast":
            loc_filter = func.lower(PotentialLead.state).in_(NORTHEAST_STATES)
        elif location == "dc":
            loc_filter = func.lower(PotentialLead.city).in_(DC_NAMES) | func.lower(
                PotentialLead.state
            ).in_(DC_NAMES)
        elif location == "midwest":
            loc_filter = func.lower(PotentialLead.state).in_(MIDWEST_STATES)
        elif location == "pacific_nw":
            loc_filter = func.lower(PotentialLead.state).in_(PACIFIC_NW_STATES)
        elif location == "las_vegas":
            loc_filter = func.lower(PotentialLead.city).in_(LAS_VEGAS_CITIES)
        elif location == "new_orleans":
            loc_filter = func.lower(PotentialLead.city).in_(NEW_ORLEANS_CITIES)
        elif location == "hawaii":
            loc_filter = (func.lower(PotentialLead.state) == "hawaii") | func.lower(
                PotentialLead.city
            ).in_(HAWAII_CITIES)

        if loc_filter is not None:
            query = query.where(loc_filter)
            count_query = count_query.where(loc_filter)

    # Sort — supports both old keys and new frontend keys
    sort_map = {
        # Original keys
        "newest": PotentialLead.created_at.desc().nullslast(),
        "oldest": PotentialLead.created_at.asc().nullslast(),
        "revenue_high": PotentialLead.revenue_opening.desc().nullslast(),
        "revenue_low": PotentialLead.revenue_opening.asc().nullslast(),
        "score_desc": PotentialLead.lead_score.desc().nullslast(),
        "score_asc": PotentialLead.lead_score.asc().nullslast(),
        "name_asc": PotentialLead.hotel_name.asc().nullslast(),
        "opening": PotentialLead.opening_date.asc().nullslast(),
        # New keys from React frontend
        "name_desc": PotentialLead.hotel_name.desc().nullslast(),
        "opening_asc": PotentialLead.opening_date.asc().nullslast(),
        "opening_desc": PotentialLead.opening_date.desc().nullslast(),
        "tier_asc": PotentialLead.brand_tier.asc().nullslast(),
        "tier_desc": PotentialLead.brand_tier.desc().nullslast(),
        "time_asc": PotentialLead.timeline_label.asc().nullslast(),
        "time_desc": PotentialLead.timeline_label.desc().nullslast(),
        "location_asc": PotentialLead.city.asc().nullslast(),
        "location_desc": PotentialLead.city.desc().nullslast(),
    }
    order_by = sort_map.get(
        sort or "score_desc", PotentialLead.lead_score.desc().nullslast()
    )

    leads, total, pages = await paginate_leads(
        db, query, count_query, page, per_page, order_by=order_by
    )
    return lead_list_response(leads, total, page, per_page, pages)


# -----------------------------------------------------------------------------
# Shortcut Lead Lists
# -----------------------------------------------------------------------------


@router.get("/leads/hot", response_model=LeadListResponse, tags=["Leads"])
async def get_hot_leads(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
):
    """Get hot leads (score >= config threshold)"""
    where = [
        PotentialLead.lead_score >= settings.hot_lead_threshold,
        PotentialLead.status == "new",
    ]
    query = select(PotentialLead).where(*where)
    count_query = select(func.count(PotentialLead.id)).where(*where)
    leads, total, pages = await paginate_leads(
        db, query, count_query, page, per_page, order_by=PotentialLead.lead_score.desc()
    )
    return lead_list_response(leads, total, page, per_page, pages)


@router.get("/leads/florida", response_model=LeadListResponse, tags=["Leads"])
async def get_florida_leads(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
):
    """Get Florida leads"""
    query = select(PotentialLead).where(PotentialLead.location_type == "florida")
    count_query = select(func.count(PotentialLead.id)).where(
        PotentialLead.location_type == "florida"
    )
    leads, total, pages = await paginate_leads(
        db,
        query,
        count_query,
        page,
        per_page,
        order_by=PotentialLead.lead_score.desc().nullslast(),
    )
    return lead_list_response(leads, total, page, per_page, pages)


@router.get("/leads/caribbean", response_model=LeadListResponse, tags=["Leads"])
async def get_caribbean_leads(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
):
    """Get Caribbean leads"""
    query = select(PotentialLead).where(PotentialLead.location_type == "caribbean")
    count_query = select(func.count(PotentialLead.id)).where(
        PotentialLead.location_type == "caribbean"
    )
    leads, total, pages = await paginate_leads(
        db,
        query,
        count_query,
        page,
        per_page,
        order_by=PotentialLead.lead_score.desc().nullslast(),
    )
    return lead_list_response(leads, total, page, per_page, pages)


# -----------------------------------------------------------------------------
# Single Lead CRUD
# -----------------------------------------------------------------------------


@router.get("/leads/{lead_id}", response_model=LeadResponse, tags=["Leads"])
async def get_lead(lead_id: int, db: AsyncSession = Depends(get_db)):
    """Get a single lead by ID"""
    result = await db.execute(select(PotentialLead).where(PotentialLead.id == lead_id))
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    return LeadResponse.model_validate(lead)


@router.post("/leads", response_model=LeadResponse, tags=["Leads"])
async def create_lead(lead_data: LeadCreate, db: AsyncSession = Depends(get_db)):
    """Create a new lead manually."""
    lead_dict = lead_data.model_dump()
    lead_dict["source_site"] = lead_dict.get("source_site") or "manual"

    result = await save_lead_to_db(lead_dict, db, commit=True)

    if result["status"] == "skipped":
        raise HTTPException(status_code=422, detail=result["reason"])
    if result["status"] in ("duplicate", "enriched"):
        raise HTTPException(
            status_code=409,
            detail=f"A lead with a similar name already exists (ID: {result['id']})",
        )

    lead = (
        await db.execute(select(PotentialLead).where(PotentialLead.id == result["id"]))
    ).scalar_one()
    logger.info(
        f"Created lead: {lead.hotel_name} (ID: {lead.id}, Score: {lead.lead_score})"
    )
    return LeadResponse.model_validate(lead)


@router.patch("/leads/{lead_id}", response_model=LeadResponse, tags=["Leads"])
async def update_lead(
    lead_id: int, updates: LeadUpdate, db: AsyncSession = Depends(get_db)
):
    """Update a lead"""
    result = await db.execute(select(PotentialLead).where(PotentialLead.id == lead_id))
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    update_data = updates.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(lead, field, value)

    lead.updated_at = local_now()
    await db.commit()
    await db.refresh(lead)
    logger.info(f"Updated lead: {lead.hotel_name} (ID: {lead.id})")

    # Auto-recalculate revenue if any revenue-relevant field was changed
    revenue_fields = {
        "room_count",
        "brand_tier",
        "brand",
        "hotel_type",
        "city",
        "state",
        "country",
    }
    if update_data.keys() & revenue_fields:
        try:
            from app.services.revenue_updater import update_lead_revenue

            await update_lead_revenue(lead_id)
            logger.info(f"Revenue recalculated for lead {lead_id} after field update")
        except Exception as e:
            logger.warning(f"Revenue recalc failed for lead {lead_id}: {e}")

    return LeadResponse.model_validate(lead)


# NOTE: All lead actions (approve, reject, restore, delete) are handled
# exclusively by /api/dashboard/leads/ routes in dashboard.py.
# The frontend uses those endpoints. Legacy /leads/{id}/approve etc. and
# /api/leads/{id}/approve have been removed to eliminate dangerous
# three-way route divergence (audit trail was missing from dashboard routes).


# ══════════════════════════════════════════════════════════════════
# GEO ENRICHMENT ROUTES
# ══════════════════════════════════════════════════════════════════


@router.post("/leads/{lead_id}/enrich-geo", tags=["Leads"])
async def enrich_lead_geo_route(
    lead_id: int,
    db: AsyncSession = Depends(get_db),
):
    """
    Trigger website discovery + geocoding for a single lead.
    Updates hotel_website, latitude, longitude in place.
    """
    from app.services.lead_geo_enrichment import enrich_lead_geo
    from sqlalchemy import update as sql_update

    result = await db.execute(select(PotentialLead).where(PotentialLead.id == lead_id))
    lead = result.scalar_one_or_none()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    geo = await enrich_lead_geo(
        hotel_name=lead.hotel_name,
        city=lead.city,
        state=lead.state,
        country=lead.country,
        brand=lead.brand,
        existing_website=lead.hotel_website,
        address=lead.address,
        zip_code=lead.zip_code,
    )
    await db.execute(
        sql_update(PotentialLead)
        .where(PotentialLead.id == lead_id)
        .values(
            latitude=geo.get("latitude") or lead.latitude,
            longitude=geo.get("longitude") or lead.longitude,
            hotel_website=geo.get("hotel_website") or lead.hotel_website,
            website_verified=geo.get("website_verified") or lead.website_verified,
        )
    )
    await db.commit()

    return {
        "id": lead_id,
        "hotel_name": lead.hotel_name,
        "latitude": geo.get("latitude"),
        "longitude": geo.get("longitude"),
        "hotel_website": geo.get("hotel_website"),
        "website_verified": geo.get("website_verified"),
    }


@router.post("/leads/bulk-enrich-geo", tags=["Leads"])
async def bulk_enrich_geo(
    db: AsyncSession = Depends(get_db),
    limit: int = Query(50, description="Max leads to process"),
    force: bool = Query(False, description="Re-enrich leads that already have coords"),
):
    """
    Backfill geocoords + websites for all existing leads that are missing them.
    Processes up to `limit` leads per call. Run multiple times to backfill all.
    """
    from app.services.lead_geo_enrichment import enrich_lead_geo
    from sqlalchemy import update as sql_update, or_

    # Find leads missing coords (or force re-enrich all)
    query = select(PotentialLead).where(
        PotentialLead.status.notin_(["rejected", "expired"])
    )
    if not force:
        query = query.where(
            or_(
                PotentialLead.latitude.is_(None),
                PotentialLead.hotel_website.is_(None),
            )
        )
    query = query.order_by(PotentialLead.lead_score.desc()).limit(limit)

    result = await db.execute(query)
    leads = result.scalars().all()

    updated = 0
    failed = 0

    for lead in leads:
        try:
            geo = await enrich_lead_geo(
                hotel_name=lead.hotel_name,
                city=lead.city,
                state=lead.state,
                country=lead.country,
                brand=lead.brand,
                existing_website=lead.hotel_website,
            )
            try:
                await db.execute(
                    sql_update(PotentialLead)
                    .where(PotentialLead.id == lead.id)
                    .values(
                        latitude=geo.get("latitude") or lead.latitude,
                        longitude=geo.get("longitude") or lead.longitude,
                        hotel_website=geo.get("hotel_website") or lead.hotel_website,
                        website_verified=geo.get("website_verified")
                        or lead.website_verified,
                    )
                )
                await db.commit()
                updated += 1
                logger.info(f"Geo enriched [{updated}/{len(leads)}]: {lead.hotel_name}")
            except Exception as db_err:
                await db.rollback()  # Reset transaction so next lead can proceed
                failed += 1
                logger.warning(
                    f"Geo enrichment DB save failed for {lead.hotel_name}: {db_err}"
                )
        except Exception as e:
            await db.rollback()
            failed += 1
            logger.warning(f"Geo enrichment failed for {lead.hotel_name}: {e}")

    return {
        "processed": len(leads),
        "updated": updated,
        "failed": failed,
        "message": "Run again to process more leads"
        if len(leads) == limit
        else "All leads processed",
    }


# ══════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════


@router.get("/leads/export", tags=["Leads"])
async def export_leads_excel(
    db: AsyncSession = Depends(get_db),
    status: Optional[str] = Query(
        None, description="Filter by status (new, approved, etc)"
    ),
    timeline: Optional[str] = Query(
        None, description="Filter by timeline label (HOT, URGENT, WARM, COOL)"
    ),
    tier: Optional[str] = Query(None, description="Filter by brand tier"),
    min_score: Optional[int] = Query(None, description="Minimum lead score"),
):
    """
    Export leads to Excel (.xlsx).
    Returns a downloadable file with all lead data, contacts, brand intel, and procurement info.
    """
    import io
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from app.models.lead_contact import LeadContact
    from app.config.brand_registry import BrandRegistry

    # ── Fetch leads ──
    q = select(PotentialLead).where(
        PotentialLead.status.notin_(["rejected", "expired"])
    )
    if status:
        statuses = [s.strip() for s in status.split(",")]
        q = q.where(PotentialLead.status.in_(statuses))
    if timeline:
        labels = [t.strip().upper() for t in timeline.split(",")]
        q = q.where(PotentialLead.timeline_label.in_(labels))
    if tier:
        q = q.where(PotentialLead.brand_tier == tier)
    if min_score:
        q = q.where(PotentialLead.lead_score >= min_score)
    q = q.order_by(PotentialLead.lead_score.desc())

    result = await db.execute(q)
    leads = result.scalars().all()

    # ── Fetch best contact per lead ──
    if leads:
        lead_ids = [lead_obj.id for lead_obj in leads]
        contacts_result = await db.execute(
            select(LeadContact)
            .where(LeadContact.lead_id.in_(lead_ids))
            .order_by(LeadContact.score.desc())
        )
        all_contacts = contacts_result.scalars().all()
        contacts_by_lead: dict[int, list] = {}
        for c in all_contacts:
            contacts_by_lead.setdefault(c.lead_id, []).append(c)
    else:
        contacts_by_lead = {}

    # ── Build workbook ──
    wb = Workbook()

    # ── Color palette ──
    NAVY = "0F1D32"
    WHITE = "FFFFFF"
    LIGHT_BG = "F8F9FA"
    URGENT_C = "DC2626"
    HOT_C = "EA580C"
    WARM_C = "D97706"
    COOL_C = "2563EB"
    TBD_C = "6B7280"
    T1_C = "D4A853"
    T2_C = "C49A3C"
    T3_C = "3E638C"
    T4_C = "6B665E"
    HIGH_C = "059669"
    MED_C = "D97706"
    LOW_C = "DC2626"

    TIMELINE_COLORS = {
        "URGENT": URGENT_C,
        "HOT": HOT_C,
        "WARM": WARM_C,
        "COOL": COOL_C,
        "TBD": TBD_C,
    }
    TIER_COLORS = {
        "tier1_ultra_luxury": T1_C,
        "tier2_luxury": T2_C,
        "tier3_upper_upscale": T3_C,
        "tier4_upscale": T4_C,
    }
    TIER_LABELS = {
        "tier1_ultra_luxury": "Ultra Luxury",
        "tier2_luxury": "Luxury",
        "tier3_upper_upscale": "Upper Upscale",
        "tier4_upscale": "Upscale",
    }
    OPP_COLORS = {"high": HIGH_C, "medium": MED_C, "low": LOW_C}

    def header_style(cell, bg=NAVY, fg=WHITE, size=10, bold=True):
        cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    def data_style(cell, bold=False, wrap=False, align="left", color=None):
        cell.font = Font(name="Arial", size=9, bold=bold, color=color or "000000")
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)

    def colored_badge(cell, text, hex_color):
        cell.value = text
        cell.font = Font(name="Arial", size=9, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=hex_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fmt_currency(val):
        if not val:
            return "—"
        if val >= 1_000_000:
            return f"${val / 1_000_000:.1f}M"
        if val >= 1_000:
            return f"${val / 1_000:.0f}K"
        return f"${val:,.0f}"

    # ════════════════════════════════════════════════════════
    # SHEET 1 — LEADS OVERVIEW
    # ════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Leads"
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False

    # Title row
    ws.merge_cells("A1:R1")
    title_cell = ws["A1"]
    title_cell.value = f"Lead Generator — Hotel Pipeline Export  ({len(leads)} leads)"
    title_cell.font = Font(name="Arial", bold=True, size=14, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    # Column headers
    HEADERS = [
        ("Score", 5),
        ("Hotel Name", 28),
        ("Brand", 18),
        ("Tier", 13),
        ("Timeline", 9),
        ("Location", 22),
        ("Opening", 12),
        ("Rooms", 7),
        ("Opening Rev.", 12),
        ("Annual Rev.", 12),
        ("Opportunity", 11),
        ("Procurement", 14),
        ("Decision Maker", 22),
        ("Website", 26),
        ("Mgmt Company", 20),
        ("Contact Name", 18),
        ("Contact Email", 24),
        ("Contact LinkedIn", 30),
    ]
    for col_idx, (label, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        header_style(cell, bg=NAVY if col_idx <= 10 else "1E3A5F")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 30

    # Data rows
    for row_idx, lead in enumerate(leads, 3):
        brand_info = BrandRegistry.lookup(lead.brand or "")
        contacts = contacts_by_lead.get(lead.id, [])
        # Best contact: primary first, then highest score
        best = next((c for c in contacts if c.is_primary), None) or (
            contacts[0] if contacts else None
        )
        # Best email contact
        email_contact = next((c for c in contacts if c.email), best)

        bg = "FFFFFF" if row_idx % 2 == 0 else LIGHT_BG
        fill = PatternFill("solid", fgColor=bg)

        def cell_val(col, value, bold=False, wrap=False, align="left", color=None):
            c = ws.cell(row=row_idx, column=col, value=value)
            data_style(c, bold=bold, wrap=wrap, align=align, color=color)
            c.fill = fill
            c.border = border
            return c

        # Score badge
        score = lead.lead_score or 0
        score_c = ws.cell(row=row_idx, column=1, value=score)
        score_color = HOT_C if score >= 70 else WARM_C if score >= 50 else COOL_C
        score_c.font = Font(name="Arial", size=9, bold=True, color=WHITE)
        score_c.fill = PatternFill("solid", fgColor=score_color)
        score_c.alignment = Alignment(horizontal="center", vertical="center")
        score_c.border = border

        cell_val(2, lead.hotel_name, bold=True)
        cell_val(3, lead.brand or "—")

        # Tier badge
        tier_text = TIER_LABELS.get(lead.brand_tier or "", lead.brand_tier or "—")
        tier_c = ws.cell(row=row_idx, column=4, value=tier_text)
        if lead.brand_tier in TIER_COLORS:
            colored_badge(tier_c, tier_text, TIER_COLORS[lead.brand_tier])
        else:
            data_style(tier_c, align="center")
        tier_c.fill = PatternFill(
            "solid", fgColor=TIER_COLORS.get(lead.brand_tier or "", "94A3B8")
        )
        tier_c.border = border

        # Timeline badge
        tl = lead.timeline_label or "TBD"
        tl_c = ws.cell(row=row_idx, column=5, value=tl)
        colored_badge(tl_c, tl, TIMELINE_COLORS.get(tl, TBD_C))
        tl_c.border = border

        location = ", ".join(filter(None, [lead.city, lead.state, lead.country]))
        cell_val(6, location or "—")
        cell_val(7, lead.opening_date or "—", align="center")
        cell_val(8, lead.room_count or "—", align="center")
        cell_val(
            9,
            fmt_currency(lead.revenue_opening),
            align="right",
            bold=bool(lead.revenue_opening),
            color=HIGH_C if lead.revenue_opening else None,
        )
        cell_val(10, fmt_currency(lead.revenue_annual), align="right")

        # Opportunity badge
        opp = brand_info.opportunity_level
        opp_c = ws.cell(row=row_idx, column=11, value=opp.upper())
        colored_badge(opp_c, opp.upper(), OPP_COLORS.get(opp, TBD_C))
        opp_c.border = border

        cell_val(12, brand_info.procurement_model.replace("_", " ").title())
        dm = (
            brand_info.pre_opening_contact_titles[:2]
            if brand_info.pre_opening_contact_titles
            else ["General Manager"]
        )
        cell_val(13, " / ".join(dm), wrap=True)

        # Website as hyperlink
        website = lead.hotel_website
        ws_cell = ws.cell(row=row_idx, column=14, value=website or "—")
        if website:
            ws_cell.hyperlink = (
                website if website.startswith("http") else f"https://{website}"
            )
            ws_cell.font = Font(
                name="Arial", size=9, color="2563EB", underline="single"
            )
        else:
            data_style(ws_cell, color="94A3B8")
        ws_cell.fill = fill
        ws_cell.border = border

        cell_val(15, lead.management_company or "—")
        cell_val(16, best.name if best else "—", bold=bool(best))
        cell_val(
            17, (email_contact.email if email_contact and email_contact.email else "—")
        )

        # LinkedIn as hyperlink
        li_url = best.linkedin if best else None
        li_cell = ws.cell(row=row_idx, column=18, value=li_url or "—")
        if li_url:
            li_cell.hyperlink = li_url
            li_cell.font = Font(
                name="Arial", size=9, color="2563EB", underline="single"
            )
        else:
            data_style(li_cell, color="94A3B8")
        li_cell.fill = fill
        li_cell.border = border

        ws.row_dimensions[row_idx].height = 18

    # Auto filter
    ws.auto_filter.ref = f"A2:R{len(leads) + 2}"

    # ════════════════════════════════════════════════════════
    # SHEET 2 — CONTACTS (all contacts, not just primary)
    # ════════════════════════════════════════════════════════
    wc = wb.create_sheet("Contacts")
    wc.sheet_view.showGridLines = False
    wc.freeze_panes = "A3"

    wc.merge_cells("A1:J1")
    tc = wc["A1"]
    tc.value = "All Contacts — Lead Generator"
    tc.font = Font(name="Arial", bold=True, size=13, color=WHITE)
    tc.fill = PatternFill("solid", fgColor=NAVY)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wc.row_dimensions[1].height = 28

    CONTACT_HEADERS = [
        ("Hotel", 26),
        ("Brand", 16),
        ("Timeline", 9),
        ("Name", 20),
        ("Title", 22),
        ("Scope", 12),
        ("Email", 26),
        ("Phone", 14),
        ("LinkedIn", 34),
        ("Score", 7),
    ]
    for ci, (label, width) in enumerate(CONTACT_HEADERS, 1):
        cell = wc.cell(row=2, column=ci, value=label)
        header_style(cell)
        wc.column_dimensions[get_column_letter(ci)].width = width
    wc.row_dimensions[2].height = 26

    crow = 3
    for lead in leads:
        contacts = contacts_by_lead.get(lead.id, [])
        if not contacts:
            continue
        tl = lead.timeline_label or "TBD"
        for c in contacts:
            bg2 = "FFFFFF" if crow % 2 == 0 else LIGHT_BG
            fill2 = PatternFill("solid", fgColor=bg2)

            def cc(col, val, bold=False, align="left", color=None):
                cell = wc.cell(row=crow, column=col, value=val)
                cell.font = Font(
                    name="Arial", size=9, bold=bold, color=color or "000000"
                )
                cell.alignment = Alignment(horizontal=align, vertical="center")
                cell.fill = fill2
                cell.border = border

            cc(1, lead.hotel_name, bold=True)
            cc(2, lead.brand or "—")

            tl_wc = wc.cell(row=crow, column=3, value=tl)
            colored_badge(tl_wc, tl, TIMELINE_COLORS.get(tl, TBD_C))
            tl_wc.fill = PatternFill("solid", fgColor=TIMELINE_COLORS.get(tl, TBD_C))
            tl_wc.border = border

            cc(4, c.name, bold=c.is_primary)
            cc(5, c.title or "—")

            scope_text = (c.scope or "").replace("_", " ").title()
            scope_color = (
                HIGH_C
                if c.scope == "hotel_specific"
                else WARM_C
                if c.scope == "chain_area"
                else TBD_C
            )
            sc = wc.cell(row=crow, column=6, value=scope_text)
            colored_badge(sc, scope_text, scope_color)
            sc.fill = PatternFill("solid", fgColor=scope_color)
            sc.border = border

            # Email
            email_cell = wc.cell(row=crow, column=7, value=c.email or "—")
            if c.email:
                email_cell.hyperlink = f"mailto:{c.email}"
                email_cell.font = Font(
                    name="Arial", size=9, color="2563EB", underline="single"
                )
            else:
                email_cell.font = Font(name="Arial", size=9, color="94A3B8")
            email_cell.fill = fill2
            email_cell.border = border
            email_cell.alignment = Alignment(horizontal="left", vertical="center")

            cc(8, c.phone or "—")

            # LinkedIn
            li_wc = wc.cell(row=crow, column=9, value=c.linkedin or "—")
            if c.linkedin:
                li_wc.hyperlink = c.linkedin
                li_wc.font = Font(
                    name="Arial", size=9, color="2563EB", underline="single"
                )
            else:
                li_wc.font = Font(name="Arial", size=9, color="94A3B8")
            li_wc.fill = fill2
            li_wc.border = border
            li_wc.alignment = Alignment(horizontal="left", vertical="center")

            cc(
                10,
                c.score or 0,
                align="center",
                color=HIGH_C
                if (c.score or 0) >= 20
                else WARM_C
                if (c.score or 0) >= 10
                else TBD_C,
            )

            wc.row_dimensions[crow].height = 17
            crow += 1

    wc.auto_filter.ref = f"A2:J{crow - 1}"

    # ════════════════════════════════════════════════════════
    # SHEET 3 — BRAND INTEL
    # ════════════════════════════════════════════════════════
    wb_sheet = wb.create_sheet("Brand Intel")
    wb_sheet.sheet_view.showGridLines = False
    wb_sheet.freeze_panes = "A3"

    wb_sheet.merge_cells("A1:G1")
    bi_title = wb_sheet["A1"]
    bi_title.value = "Brand Procurement Intelligence — Lead Generator"
    bi_title.font = Font(name="Arial", bold=True, size=13, color=WHITE)
    bi_title.fill = PatternFill("solid", fgColor=NAVY)
    bi_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wb_sheet.row_dimensions[1].height = 28

    BI_HEADERS = [
        ("Brand", 20),
        ("Parent Company", 20),
        ("Operating Model", 16),
        ("Procurement", 18),
        ("Uniform Freedom", 13),
        ("Opportunity", 11),
        ("Pre-Opening Contact", 30),
    ]
    for bi, (label, width) in enumerate(BI_HEADERS, 1):
        cell = wb_sheet.cell(row=2, column=bi, value=label)
        header_style(cell)
        wb_sheet.column_dimensions[get_column_letter(bi)].width = width
    wb_sheet.row_dimensions[2].height = 26

    # Get unique brands from leads
    seen_brands = {}
    for lead in leads:
        b = (lead.brand or "").strip()
        if b and b not in seen_brands:
            seen_brands[b] = BrandRegistry.lookup(b)

    for bi_row, (brand_name, info) in enumerate(sorted(seen_brands.items()), 3):
        bg3 = "FFFFFF" if bi_row % 2 == 0 else LIGHT_BG
        fill3 = PatternFill("solid", fgColor=bg3)

        def bc(col, val, align="left", color=None, bold=False):
            cell = wb_sheet.cell(row=bi_row, column=col, value=val)
            cell.font = Font(name="Arial", size=9, bold=bold, color=color or "000000")
            cell.alignment = Alignment(
                horizontal=align, vertical="center", wrap_text=True
            )
            cell.fill = fill3
            cell.border = border

        bc(1, brand_name, bold=True)
        bc(2, info.parent_company)
        bc(3, info.operating_model.replace("_", " ").title(), align="center")
        bc(4, info.procurement_model.replace("_", " ").title())

        freedom_c = wb_sheet.cell(
            row=bi_row, column=5, value=info.uniform_freedom.upper()
        )
        fc = (
            HIGH_C
            if info.uniform_freedom == "high"
            else WARM_C
            if info.uniform_freedom == "medium"
            else LOW_C
        )
        colored_badge(freedom_c, info.uniform_freedom.upper(), fc)
        freedom_c.fill = PatternFill("solid", fgColor=fc)
        freedom_c.border = border

        opp_bc = wb_sheet.cell(
            row=bi_row, column=6, value=info.opportunity_level.upper()
        )
        oc = OPP_COLORS.get(info.opportunity_level, TBD_C)
        colored_badge(opp_bc, info.opportunity_level.upper(), oc)
        opp_bc.fill = PatternFill("solid", fgColor=oc)
        opp_bc.border = border

        bc(7, " / ".join(info.pre_opening_contact_titles[:3]))
        wb_sheet.row_dimensions[bi_row].height = 18

    # ── Stream response ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import date

    filename = f"leads_export_{date.today().isoformat()}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
